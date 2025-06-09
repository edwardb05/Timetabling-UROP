# %%
# %pip install pandas openpyxl

# %%
import pandas as pd
from openpyxl import load_workbook

# %% [markdown]
# Some example data of what an output dictionary could look like with it being exam name and then day then slot (morning/afternoon), this only has 4 days and a few exams

# %%
exams_timetabled = {
    'Math': (0, 0),        # Monday, Morning
    'Physics': (1, 0),     # Tuesday, Morning
    'Chemistry': (1, 1),   # Monday, Afternoon
    'History': (2, 0),     # Wednesday, Morning
    'Geography': (1, 1),   # Tuesday, Afternoon
    'PE': (3, 0),          # Thursday, Morning
    'Music': (2, 1),       # Wednesday, Afternoon
}


# %%
days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday',]
slots = ['Morning', 'Afternoon']


# Group exams by day and slot
data = {}
for exam, (d, s) in exams_timetabled.items():
    day = days[d]
    slot = slots[s]
    data.setdefault(day, {}).setdefault(slot, []).append(exam)

rows = []
for day in days:
    for slot in slots:
        exams_list = data.get(day, {}).get(slot, [])
        exams_str = ', '.join(exams_list) if exams_list else ''
        rows.append([day, slot, exams_str])


df = pd.DataFrame(rows, columns=['Day', 'Time', 'Exams'])

# Save DataFrame to Excel
filename = 'exam_schedule_merged.xlsx'
df.to_excel(filename, index=False)

# Now open with openpyxl to merge cells
wb = load_workbook(filename)
ws = wb.active

def merge_cells_for_col(col_idx):
    current_val = None
    start_row = 2  # skip header
    max_row = ws.max_row + 1
    for row in range(2, max_row):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value != current_val:
            # Merge the previous block if >1 row
            if current_val is not None and row - start_row > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx,
                               end_row=row - 1, end_column=col_idx)
            current_val = cell.value
            start_row = row
    # Merge last block
    if current_val is not None and max_row - start_row > 0:
        ws.merge_cells(start_row=start_row, start_column=col_idx,
                       end_row=max_row - 1, end_column=col_idx)

# Merge Day column (A=1)
merge_cells_for_col(1)
# Merge Time column (B=2)
merge_cells_for_col(2)

wb.save(filename)
print(f"Excel file '{filename}' created with merged cells.")



