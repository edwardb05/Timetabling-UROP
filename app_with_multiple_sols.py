import streamlit as st
import pandas as pd
import numpy as np
from ortools.sat.python import cp_model
from rapidfuzz import process, fuzz
from collections import defaultdict
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re
from dateutil.parser import parse
import time
import logging
from openpyxl.styles import PatternFill, Alignment
import threading
import time
import streamlit.components.v1 as components

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

st.set_page_config(page_title="Exam Timetabling System", layout="wide")

st.title("Exam Timetabling System")
st.markdown("""
This app helps create an optimal exam timetable while considering various constraints like:
- Student exam conflicts
- Module leader preferences
- Special accommodations for students with extra time
- Fixed exam dates
- Bank holidays
- Room assignments
- AEA requirements
""")

# File upload section
st.header("Upload Required Files")
col1, col2, col3 = st.columns(3)

with col1:
    student_file = st.file_uploader("Upload Student List", type=['xlsx'])
with col2:
    module_file = st.file_uploader("Upload Module List", type=['xlsx'])
with col3:
    dates_file = st.file_uploader("Upload Useful Dates", type=['xlsx'])

# Parameters section
st.header("Timetabling Parameters")
col1, col2 = st.columns(2)

with col1:
    num_days = st.number_input("Number of Days for Exam Period", min_value=1, max_value=30, value=20)
    max_exams_2days = st.number_input("Maximum Exams in 2-Day Window", min_value=1, max_value=5, value=3)
    max_exams_5days = st.number_input("Maximum Exams in 5-Day Window", min_value=1, max_value=10, value=4)

with col2:
    spread_penalty = st.slider("Module leaders spread out Penalty Weight", min_value=0, max_value=10, value=5)
    unused_penalty = st.slider("Unused seats Penalty Weight", min_value=0, max_value=10, value=5)
    extra_time_penalty = st.slider("Extra Time Student Penalty Weight", min_value=0, max_value=10, value=5)

# Core modules list
Core_modules = ["MECH70001 Nuclear Thermal Hydraulics",
                "MECH60004/MECH70042 Introduction to Nuclear Energy A/B",
                "MECH70002 Nuclear Reactor Physics",
                "MECH70008 Mechanical Transmissions Technology",
                "MECH70006 Metal Processing Technology",
                "MECH70021Aircraft Engine Technology",
                "MECH70003 Future Clean Transport Technology",
                "MECH60015/70030 PEN3/AME"]


# Fixed modules dictionary
Fixed_modules = {"BUSI60039 Business Strategy" :[1,1],
                 "BUSI60046 Project Management":[2,1],
                 "ME-ELEC70098 Optimisation":[3,0],
                 "MECH70001 Nuclear Thermal Hydraulics":[3,0],
                 "BUSI60040/BUSI60043 Corporate Finance Online/Finance & Financial Management":[3,1],
                 "MECH60004/MECH70042 Introduction to Nuclear Energy A/B":[4,0],
                 "ME-ELEC70022 Modelling and Control of Multi-body Mechanical Systems":[4,0],
                 "MATE97022 Nuclear Materials 1":[4,0],
                 "ME-MATE70029 Nuclear Fusion":[9,0],
                 "MECH70002 Nuclear Reactor Physics":[10,0],
                 "ME-ELEC70076 Sustainable Electrical Systems":[10,0],
                 "ME ELEC70066 Applied Advanced Optimisation":[10,0],
                 "MECH70020 Combustion, Safety and Fire Dynamics":[11,0],
                 "BIOE70016 Human Neuromechanical Control and Learning":[11,0],
                 "CENG60013 Nuclear Chemical Engineering":[11,0],
                 "MECH70008 Mechanical Transmissions Technology":[17,1],
                 "MECH70006 Metal Processing Technology":[17,1],
                 "MECH70021Aircraft Engine Technology":[17,1],
                 "MECH70003 Future Clean Transport Technology":[17,1],
                 "MECH60015/70030 PEN3/AME":[18,1]}

# Room dictionary with capacities and features
rooms = {
    'CAGB 203': [["Computer", "SEQ","AEA"], 65],
    'CAGB 309': [["SEQ"], 54],
    'CAGB 659-652': [["SEQ"], 75],
    'CAGB 747-748': [["SEQ"], 36],
    'CAGB 749-752': [["SEQ"], 75],
    'CAGB 761': [["Computer"], 25],
    'CAGB 762': [["Computer"], 25],
    'SKEM 208': [["Computer"], 35],
    'SKEM 317': [["Computer"], 20],
    'CAGB 320-321': [["AEA"], 10],
    'CAGB 305': [["AEA"], 4],
    'CAGB 349': [["AEA"], 2],
    'CAGB 311': [["AEA"], 1],
    'CAGB 765': [["AEA"], 10],
    'CAGB 527': [["AEA"], 2]
}

# No exam dates (weekends and last Friday morning)
no_exam_dates = [
    [5,0], [5,1], [6,0], [6,1],  # First weekend
    [12,0], [12,1], [13,0], [13,1],  # Second weekend
    [18,0], [19,0], [19,1], [20,0], [20,1]  # Last Friday and weekend
]

def ordinal(n):
    # Returns ordinal string for an integer n, e.g. 1 -> 1st, 2 -> 2nd
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    else:
        return f"{n}{['th','st','nd','rd','th','th','th','th','th','th'][n % 10]}"

def validate_student_list(df):
    """Validate the student list Excel file format and content."""
    errors = []
    
    # Check if file has enough rows (header + at least one student)
    if len(df) < 3:
        errors.append("Student list must have at least 3 rows (header + students)")
        return errors
    
    # Check if required columns exist (CID and AEA status)
    if df.iloc[0, 0] != "CID" or df.iloc[0, 3] != "Additional Exam Arrangements AEA":
        errors.append(f"Student list must have 'CID' instead of {df.iloc[0, 0]} in column A and 'AEA' instead of {df.iloc[0, 3]}")
        return errors
    
    # Check if exam columns exist (starting from column J)
    exam_columns = df.iloc[0, 9:].dropna()
    if len(exam_columns) == 0:
        errors.append("No exam columns found starting from column J")
        return errors
    
    # Check for valid exam indicators (x, a, b)
    student_rows = df.iloc[2:, :]
    for idx, row in student_rows.iterrows():
        cid = row[0]
        if pd.isna(cid):
            errors.append(f"Missing CID in row {idx + 3}")
            continue
            
        # Check exam indicators
        for col_idx, exam_name in enumerate(exam_columns, start=9):
            value = str(row[col_idx]).strip().lower()
            if value not in ['x', 'a', 'b', 'nan']:
                errors.append(f"Invalid exam indicator '{value}' for student {cid} in exam {exam_name}")
    
    return errors

def validate_module_list(df):
    """Validate the module list Excel file format and content."""
    errors = []
    
    # Check if file has enough rows
    if len(df) < 2:
        errors.append("Module list must have at least 2 rows")
        return errors
    
    # Check required columns
    required_cols = ['Banner Code (New CR)', 'Module Name', 'Module Leader (lecturer 1)']
    for col in required_cols:
        if col not in df.columns:
            errors.append(f"Missing required column: {col}")
    

    return errors

def validate_useful_dates(wb):
    """Validate the useful dates Excel file format and content."""
    errors = []
    
    if not wb:
        errors.append("Could not open useful dates file")
        return errors
    
    ws = wb.active
    
    # Check if file has bank holidays section
    found_bank_holidays = False
    row = 5
    while True:
        name = ws[f"F{row}"].value
        if name is None or "Term Dates" in str(name):
            break
        if "Bank Holiday" in str(name):
            found_bank_holidays = True
            break
        row += 1
    
    if not found_bank_holidays:
        errors.append("Could not find bank holidays section in useful dates file")
    
    # Check if Summer Term section exists
    found_summer_term = False
    row = 5
    while row < ws.max_row:
        cell_value = ws[f"F{row}"].value
        if cell_value and "Summer Term" in str(cell_value):
            found_summer_term = True
            break
        row += 1
    
    if not found_summer_term:
        errors.append("Could not find Summer Term section in useful dates file")
    
    return errors


def process_files():
    """Process uploaded files and return processed data."""
    if not all([student_file, module_file, dates_file]):
        st.error("Please upload all required files")
        return None, None, None
    
    try:
        # Read files
        student_df = pd.read_excel(student_file, header=None)  # No header for student list
        module_df = pd.read_excel(module_file, sheet_name=1, header=1)  # Sheet 1, header row 1
        dates_wb = load_workbook(dates_file)
        
        # Validate files
        student_errors = validate_student_list(student_df)
        if student_errors:
            st.error("Student list errors:\n" + "\n".join(student_errors))
            return None, None, None
            
        module_errors = validate_module_list(module_df)
        if module_errors:
            st.error("Module list errors:\n" + "\n".join(module_errors))
            return None, None, None
            
        dates_errors = validate_useful_dates(dates_wb)
        if dates_errors:
            st.error("Useful dates errors:\n" + "\n".join(dates_errors))
            return None, None, None
        exams = student_df.iloc[0, 9:].dropna().tolist()
    # Get the range of rows containing student data (from row 3 onward)
        student_rows = student_df.iloc[2:, :] 
        student_exams = {}
        for _, row in student_rows.iterrows():
            cid = row[0]  # Column A = student CID
            exams_taken = []

            for col_idx, exam_name in enumerate(exams, start=9):  # Column J = index 9
                if str(row[col_idx]).strip().lower() == 'x' or str(row[col_idx]).strip().lower() == 'a'  or str(row[col_idx]).strip().lower() == 'b' :  # Check for 'x' or 'a' or 'b' to indicate they take this course (case-insensitive)
                    exams_taken.append(exam_name)

            student_exams[cid] = exams_taken
        for student in student_exams:
            for exam in student_exams[student]:
                if exam in Core_modules:
                    for other_exam in Fixed_modules:
                        if other_exam in student_exams[student]:
                            if exam != other_exam and Fixed_modules[exam][0] == Fixed_modules[other_exam][0]:
                                st.error(f"Core module {exam} conflicts with fixed module {other_exam} on the same day for student {student} so model will be infeasible")

                        
        return student_df, module_df, dates_wb
        
    except Exception as e:
        st.error(f"Error processing files: {str(e)}")
        return None, None, None


def create_timetable(students_df, leaders_df, wb,max_exams_2days, max_exams_5days):

    # Extract exam names from row 0, starting from column J (index 9)
    exams = students_df.iloc[0, 9:].dropna().tolist()
    # Get the range of rows containing student data (from row 3 onward)
    student_rows = students_df.iloc[2:, :]  # row index 3 and onward
    #Total student count for AEA and non AEA

    # Process bank holidays and create no_exam_dates

    ws = wb.active
    bank_holidays = []
    row = 5
    while True:
        name = ws[f"F{row}"].value
        date_cell = ws[f"G{row}"].value
        if name is None or "Term Dates" in str(name):
            break
        if isinstance(date_cell, datetime):
            bank_holidays.append((str(name).strip(), date_cell.date()))
        row += 1
    
    # Find Summer Term start date
    summer_start = None
    while row < ws.max_row:
        cell_value = ws[f"F{row}"].value
        if cell_value and "Summer Term" in str(cell_value):
            term_range = ws[f"F{row + 1}"].value
            if term_range:
                try:
                    start_part = term_range.split("to")[0].strip()
                    start_str = re.sub(r"^\w+\s+", "", start_part)
                    year_match = re.search(r"\b\d{4}\b", term_range)
                    if year_match:
                        start_str += f" {year_match.group(0)}"
                    else:
                        st.error("Year not found in date range.")
                        return None
                    summer_start = parse(start_str, dayfirst=True).date()
                except Exception as e:
                    st.error(f"Could not parse Summer Term start: {term_range}")
                    return None
            break
        row += 1

    if not summer_start:
        st.error("Summer Term start date not found")
        return None

    # Find first Monday
    first_monday = summer_start

    while first_monday.weekday() != 0:
        first_monday += timedelta(days=1)
    for name, bh_date in bank_holidays:
        delta = (bh_date - first_monday).days
        if 0 <= delta <= 20:
            no_exam_dates.append([delta, 0])
            no_exam_dates.append([delta, 1])

    student_exams = {}

    for _, row in student_rows.iterrows():
        cid = row[0]  # Column A = student CID
        exams_taken = []

        for col_idx, exam_name in enumerate(exams, start=9):  # Column J = index 9
            if str(row[col_idx]).strip().lower() == 'x' or str(row[col_idx]).strip().lower() == 'a'  or str(row[col_idx]).strip().lower() == 'b' :  # Check for 'x' or 'a' or 'b' to indicate they take this course (case-insensitive)
                exams_taken.append(exam_name)

        student_exams[cid] = exams_taken

        # Get the range of rows containing student data (from row 3 onward)
    student_rows = students_df.iloc[2:, :]  # row index 3 and onward

    days = []
    for i in range(21):
        date = first_monday + timedelta(days=i)
        day_str = date.strftime("%A ") + ordinal(date.day) + date.strftime(" %B")
        days.append(day_str)

    # Create a boolean mask for column D where value is not "#N/A" (after stripping whitespace)
    valid_aea_mask = (
        student_rows.iloc[:, 3].notna() &
        (student_rows.iloc[:, 3].astype(str).str.strip() != "#N/A")
    )

    # Get the CIDs (from column 0) where AEA condition is met
    AEA = student_rows.loc[valid_aea_mask, student_rows.columns[0]].tolist()
    standardized_names = exams

    leader_courses = defaultdict(list)

    # Loop through rows in the module list
    for _, row in leaders_df.iterrows():
        leader = row['Module Leader (lecturer 1)']
        name = row['Module Name']
        code = row['Banner Code (New CR)']   # module leader

        # Skip if any required field is missing
        if pd.isna(code) or pd.isna(name) or pd.isna(leader):
            continue

        if leader == "n/a":
            continue

        # Combine code and name
        combined_name = f"{code} {name}"

        

        # Fuzzy match to standardized names
        best_match, score, _ = process.extractOne(
            combined_name, standardized_names, scorer=fuzz.token_sort_ratio
        )
        if score >= 70:
            if best_match not in leader_courses[leader]:
                leader_courses[leader].append(best_match)
            else:
                st.write(f"⚠️ Duplicate match skipped for '{combined_name}': '{best_match}' is already listed for {leader}.")
        else:
            st.write(f"⚠️ Low confidence match for '{combined_name}' (best: '{best_match}', score: {score}).")

    exam_counts = defaultdict(lambda: [0, 0])

    for cid, exams_taken in student_exams.items():
        if cid in AEA:
            for exam in exams_taken:
                exam_counts[exam][0] += 1
        else:
            for exam in exams_taken:
                exam_counts[exam][1] += 1
    # Process AEA students
    # Convert to normal dict if desired
    # Process extra time students
    extra_time_students_25 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("15min/hour", "25% extra time"))].iloc[:, 0].tolist()
    # Make a list of students with 50% extra time to ensure they dont have more than on exam a day
    extra_time_students_50 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("30min/hour", "50% extra time"))].iloc[:, 0].tolist()





    # Initialize model
    model = cp_model.CpModel()
    slots = [0, 1]
    num_slots = len(slots)
    num_days = len(days)

    # Variables: exam_day and exam_slot
    exam_day = {}
    exam_slot = {}
    for exam in exams:
        exam_day[exam] = model.NewIntVar(0, num_days - 1, f'{exam}_day')
        exam_slot[exam] = model.NewIntVar(0, num_slots - 1, f'{exam}_slot')

    # Variable to put exams in rooms
    exam_room = {}
    for exam in set().union(*student_exams.values()):
        for room in rooms:
            exam_room[(exam, room)] = model.NewBoolVar(f'{exam}_in_{room.replace(" ", "_")}')

    # 1. Core modules can not have multiple exams on that day
    for student, exs in student_exams.items():
        core_mods = [exam for exam in exs if exam in Core_modules]
        other_mods = [exam for exam in exs if exam not in Core_modules]
        for exam in core_mods:
            for other in other_mods:
                model.Add(exam_day[exam] != exam_day[other])

    # 2. Fixed modules day and slot assignment
    for exam, (day_fixed, slot_fixed) in Fixed_modules.items():
        model.Add(exam_day[exam] == day_fixed)
        model.Add(exam_slot[exam] == slot_fixed)

    # 3. Forbidden exam day-slot assignments
    for exam in exams:
        for day, slot in no_exam_dates:
            model.AddForbiddenAssignments([exam_day[exam], exam_slot[exam]], [(day, slot)])
    # 4. Max 3 exams in any 2-day window per student
    for student, ex in student_exams.items():
        for d in range(num_days - 1):
            exams_in_2_days = []
            for exam in ex:
                is_on_d = model.NewBoolVar(f'{student}_{exam}_on_day_{d}')
                is_on_d1 = model.NewBoolVar(f'{student}_{exam}_on_day_{d+1}')
                is_on_either = model.NewBoolVar(f'{student}_{exam}_on_day_{d}_or_{d+1}')

                model.Add(exam_day[exam] == d).OnlyEnforceIf(is_on_d)
                model.Add(exam_day[exam] != d).OnlyEnforceIf(is_on_d.Not())

                model.Add(exam_day[exam] == d + 1).OnlyEnforceIf(is_on_d1)
                model.Add(exam_day[exam] != d + 1).OnlyEnforceIf(is_on_d1.Not())

                model.AddBoolOr([is_on_d, is_on_d1]).OnlyEnforceIf(is_on_either)
                model.AddBoolAnd([is_on_d.Not(), is_on_d1.Not()]).OnlyEnforceIf(is_on_either.Not())

                exams_in_2_days.append(is_on_either)

            model.Add(sum(exams_in_2_days) <= max_exams_2days)

    # 5. Max 4 exams in any 5-day sliding window per student
    for student, exs in student_exams.items():
        for start_day in range(num_days - 4):
            exams_in_window = []
            for exam in exs:
                in_window = model.NewBoolVar(f'{student}_{exam}_in_day_{start_day}_to_{start_day + 4}')

                model.AddLinearConstraint(exam_day[exam], start_day, start_day + 4).OnlyEnforceIf(in_window)

                before_window = model.NewBoolVar(f'{student}_{exam}_before_{start_day}')
                after_window = model.NewBoolVar(f'{student}_{exam}_after_{start_day + 4}')

                model.Add(exam_day[exam] < start_day).OnlyEnforceIf(before_window)
                model.Add(exam_day[exam] >= start_day).OnlyEnforceIf(before_window.Not())

                model.Add(exam_day[exam] > start_day + 4).OnlyEnforceIf(after_window)
                model.Add(exam_day[exam] <= start_day + 4).OnlyEnforceIf(after_window.Not())

                model.AddBoolOr([before_window, after_window]).OnlyEnforceIf(in_window.Not())

                exams_in_window.append(in_window)

            model.Add(sum(exams_in_window) <= max_exams_5days)

    # 6. At most 1 exam in week 3 (days 13 to 20) per module leader
    for leader, leader_exams in leader_courses.items():
        exams_in_week3 = []
        for exam in leader_exams:
            in_week3 = model.NewBoolVar(f'{exam}_in_week3')

            model.AddLinearConstraint(exam_day[exam], 13, 20).OnlyEnforceIf(in_week3)

            before_week3 = model.NewBoolVar(f'{exam}_before_week3')
            after_week3 = model.NewBoolVar(f'{exam}_after_week3')

            model.Add(exam_day[exam] < 13).OnlyEnforceIf(before_week3)
            model.Add(exam_day[exam] >= 13).OnlyEnforceIf(before_week3.Not())

            model.Add(exam_day[exam] > 20).OnlyEnforceIf(after_week3)
            model.Add(exam_day[exam] <= 20).OnlyEnforceIf(after_week3.Not())

            model.AddBoolOr([before_week3, after_week3]).OnlyEnforceIf(in_week3.Not())

            exams_in_week3.append(in_week3)

        model.Add(sum(exams_in_week3) <= 1)

    # 7. Extra time 50% students: max 1 exam per day
    for student in extra_time_students_50:
        for day in range(num_days):
            exams_on_day = []
            for exam in student_exams[student]:
                is_on_day = model.NewBoolVar(f'{student}_{exam}_on_day_{day}')
                model.Add(exam_day[exam] == day).OnlyEnforceIf(is_on_day)
                model.Add(exam_day[exam] != day).OnlyEnforceIf(is_on_day.Not())
                exams_on_day.append(is_on_day)
            model.Add(sum(exams_on_day) <= 1)

    # Soft constraint that extra time students with<= 25% should only have one a day
    extra_time_gr8er_1_day = []

    for student in extra_time_students_25:
        for day in range(num_days):
            exams_on_day = []
            for exam in student_exams[student]:
                is_on_day = model.NewBoolVar(f'{student}_{exam}_on_day_{day}')
                model.Add(exam_day[exam] == day).OnlyEnforceIf(is_on_day)
                model.Add(exam_day[exam] != day).OnlyEnforceIf(is_on_day.Not())
                exams_on_day.append(is_on_day)

            # Total number of exams student has that day
            num_exams = model.NewIntVar(0, len(exams_on_day), f'{student}_num_exams_day_{day}')
            model.Add(num_exams == sum(exams_on_day))

            # Create reified bool for "more than one exam"
            has_multiple_exams = model.NewBoolVar(f'{student}_more_than_one_exam_day_{day}')
            model.Add(num_exams >= 2).OnlyEnforceIf(has_multiple_exams)
            model.Add(num_exams < 2).OnlyEnforceIf(has_multiple_exams.Not())

            # Link this to a penalty variable
            penalty = model.NewIntVar(0, 1, f'{student}_penalty_day_{day}')
            model.Add(penalty == 1).OnlyEnforceIf(has_multiple_exams)
            model.Add(penalty == 0).OnlyEnforceIf(has_multiple_exams.Not())

            extra_time_gr8er_1_day.append(penalty)

    # Soft constraint that course leaders modules should be spread out
    spread_penalties = []
    for leader in leader_courses:
        mods = leader_courses[leader]
        for i in range(len(mods)):
            for j in range(i+1, len(mods)):
                m1 = mods[i]
                m2 = mods[j]

                # Calculate absolute day difference
                diff = model.NewIntVar(-21, 21, f'{m1}_{m2}_diff')
                abs_diff = model.NewIntVar(0, 21, f'{m1}_{m2}_abs_diff')
                model.Add(diff == exam_day[m1] - exam_day[m2])
                model.AddAbsEquality(abs_diff, diff)

                # Create penalty variable
                close_penalty = model.NewIntVar(0, 5, f'{m1}_{m2}_penalty')

                # Create Boolean conditions
                is_gap_3 = model.NewBoolVar(f'{m1}_{m2}_gap3')
                is_gap_2 = model.NewBoolVar(f'{m1}_{m2}_gap2')
                is_gap_1 = model.NewBoolVar(f'{m1}_{m2}_gap1')
                is_gap_0 = model.NewBoolVar(f'{m1}_{m2}_gap0')

                # Set the true condition
                model.Add(abs_diff == 3).OnlyEnforceIf(is_gap_3)
                model.Add(abs_diff != 3).OnlyEnforceIf(is_gap_3.Not())

                model.Add(abs_diff == 2).OnlyEnforceIf(is_gap_2)
                model.Add(abs_diff != 2).OnlyEnforceIf(is_gap_2.Not())

                model.Add(abs_diff == 1).OnlyEnforceIf(is_gap_1)
                model.Add(abs_diff != 1).OnlyEnforceIf(is_gap_1.Not())

                model.Add(abs_diff == 0).OnlyEnforceIf(is_gap_0)
                model.Add(abs_diff != 0).OnlyEnforceIf(is_gap_0.Not())

                # Assign penalty values based gap
                model.Add(close_penalty == 1).OnlyEnforceIf(is_gap_3)
                model.Add(close_penalty == 3).OnlyEnforceIf(is_gap_2)
                model.Add(close_penalty == 4).OnlyEnforceIf(is_gap_1)
                model.Add(close_penalty == 5).OnlyEnforceIf(is_gap_0)

                # No penalty if gap ≥ 4 and not equal to 0–3
                model.Add(close_penalty == 0).OnlyEnforceIf(
                    is_gap_3.Not(), is_gap_2.Not(), is_gap_1.Not(), is_gap_0.Not()
                )

                spread_penalties.append(close_penalty)

    # Room constraints
    unuseds = []
    for exam in exams:
       		# Calculate capacity's for each room
        AEA_capacity = sum(
            rooms[room][1] * exam_room[(exam, room)]
            for room in rooms if "AEA" in rooms[room][0]
        )

        SEQ_capacity = sum(
            rooms[room][1] * exam_room[(exam, room)]
            for room in rooms if "SEQ" in rooms[room][0]
        )


        AEA_students = exam_counts[exam][0]
        SEQ_students = exam_counts[exam][1]

        # Add Constraint
        model.Add(AEA_capacity >= AEA_students)
        model.Add(SEQ_capacity >= SEQ_students)

        # Calculate unused seats
        unused_AEA = model.NewIntVar(0, 10000, f'{exam}_unused_AEA')
        unused_SEQ = model.NewIntVar(0, 10000, f'{exam}_unused_SEQ')

        model.Add(unused_AEA == AEA_capacity - AEA_students)
        model.Add(unused_SEQ == SEQ_capacity - SEQ_students)

            # Calculate total unused seats
        total_unused = model.NewIntVar(0, 20000, f'{exam}_total_unused')
        model.Add(total_unused == unused_AEA + unused_SEQ)

            # add to list unused
        unuseds.append(total_unused)

    # Room time conflicts
    for d in range(num_days):
        for s in range(num_slots):
            for room in rooms:
                exams_in_room_time = []
                for exam in set().union(*student_exams.values()):
                    # Only consider exams that *can* be scheduled in this day and slot
                    exam_at_day = model.NewBoolVar(f'{exam}_on_day_{d}')
                    model.Add(exam_day[exam] == d).OnlyEnforceIf(exam_at_day)
                    model.Add(exam_day[exam] != d).OnlyEnforceIf(exam_at_day.Not())

                    exam_at_slot = model.NewBoolVar(f'{exam}_on_slot_{s}')
                    model.Add(exam_slot[exam] == s).OnlyEnforceIf(exam_at_slot)
                    model.Add(exam_slot[exam] != s).OnlyEnforceIf(exam_at_slot.Not())

                    exam_at_time = model.NewBoolVar(f'{exam}_on_{d}_{s}')
                    model.AddBoolAnd([exam_at_day, exam_at_slot]).OnlyEnforceIf(exam_at_time)
                    model.AddBoolOr([exam_at_day.Not(), exam_at_slot.Not()]).OnlyEnforceIf(exam_at_time.Not())

                    # Now combine with room assignment
                    assigned_and_scheduled = model.NewBoolVar(f'{exam}_in_{room}_at_{d}_{s}')
                    model.AddBoolAnd([exam_room[(exam, room)], exam_at_time]).OnlyEnforceIf(assigned_and_scheduled)
                    model.AddBoolOr([exam_room[(exam, room)].Not(), exam_at_time.Not()]).OnlyEnforceIf(assigned_and_scheduled.Not())

                    exams_in_room_time.append(assigned_and_scheduled)

                # Add AtMostOne constraint: only one exam can be assigned to this room at this time
                model.AddAtMostOne(exams_in_room_time)

        # Each exam must have at least one room
    for exam in exams:
            model.Add(sum(exam_room[(exam, room)] for room in rooms) >= 1)

    model.Minimize(sum(spread_penalties*spread_penalty + extra_time_gr8er_1_day*extra_time_penalty+unuseds*unused_penalty))
    class ExamScheduleCollector(cp_model.CpSolverSolutionCallback):
            def __init__(self, exam_day, exam_slot, exam_room, exams, rooms, leader_courses, days, slots, max_solutions=10):
                cp_model.CpSolverSolutionCallback.__init__(self)
                self.exam_day = exam_day
                self.exam_slot = exam_slot
                self.exam_room = exam_room
                self.exams = exams
                self.rooms = rooms
                self.leader_courses = leader_courses
                self.days = days
                self.slots = slots
                self.spread_penalties = spread_penalties or []
                self.soft_penalties = extra_time_gr8er_1_day or []
                self.unuseds = unuseds or []
                self.solutions = []
                self.max_solutions = max_solutions

            def on_solution_callback(self):
                schedule = {}
                for exam in self.exams:
                    d = self.Value(self.exam_day[exam])
                    s = self.Value(self.exam_slot[exam])
                    assigned_rooms = [room for room in self.rooms if self.Value(self.exam_room[(exam, room)]) == 1]
                    try:
                        leader = [name for name, exams in self.leader_courses.items() if exam in exams][0]
                    except IndexError:
                        leader = "unknown"
                    schedule[exam] = (d, s, assigned_rooms)

            # Calculate total penalty
                total_penalty = sum(self.Value(v) for v in self.spread_penalties + self.soft_penalties + self.unuseds)

            # Store as (schedule, penalty)
                self.solutions.append((schedule, total_penalty))

                if len(self.solutions) >= self.max_solutions:
                    self.StopSearch()

        # Usage:

    solver = cp_model.CpSolver()
    solver.parameters.enumerate_all_solutions = True

    collector = ExamScheduleCollector(
            exam_day, exam_slot, exam_room,
            exams, rooms, leader_courses, days, slots,
            max_solutions=100
        )

    status = solver.Solve(model, collector)

    if status == cp_model.INFEASIBLE:
            st.error('Infeasible model.')
    elif len(collector.solutions) == 0:
            st.error('No solutions found.')
    else:
            st.write(f"Found {len(collector.solutions)} solutions.")

            # Sort by penalty (ascending)
            sorted_solutions = sorted(collector.solutions, key=lambda tup: tup[1])  # tup = (schedule, penalty)
            return sorted_solutions, days,exam_counts

def generate_excel(sorted_solutions, days,exam_counts):
    
    for i, (exams_timetabled,Penalties),  in enumerate(sorted_solutions):
        # ------------ BUILD data dictionary ------------
        # data[day][slot] = list of (exam_name, rooms)
        data = {}
        for exam, (d, s, room) in exams_timetabled.items():
            day = days[d]
            slot = s  # keep slot as int 0 or 1
            data.setdefault(day, {}).setdefault(slot, []).append((exam, room))

        # ------------ BUILD rows and row_meta ------------
        rows = []
        row_meta = []  # will store tuples (day_idx, slot_idx) to track rows for merges and coloring

        for d_idx, day_name in enumerate(days):
            for s_idx, slot_name in enumerate(['Morning', 'Afternoon']):
                exams_list = data.get(day_name, {}).get(s_idx, [])
                if not exams_list:
                    # No exams this slot — add empty row with '' students and empty room
                    rows.append([day_name, slot_name, '', '', ''])
                    row_meta.append((d_idx, s_idx))
                else:
                    for exam_name, room in exams_list:
                        room_str = ', '.join(room)
                        total_students = sum(exam_counts.get(exam_name, [0, 0]))  # safe get
                        rows.append([day_name, slot_name, exam_name, total_students, room_str])
                        row_meta.append((d_idx, s_idx))

        # ------------ SAVE to Excel ------------
        df = pd.DataFrame(rows, columns=['Date', 'Time', 'Exam', 'Total No of Students', 'Room'])
        filename = f'exam_schedule_merged{i}.xlsx'
        df.to_excel(filename, index=False)

        # ------------ LOAD workbook and worksheet ------------
        wb = load_workbook(filename)
        ws = wb.active

        # ------------ FUNCTION to merge vertical cells ------------
        def merge_vertical(col, key_fn):
            start = 2
            last_key = key_fn(start)
            for r in range(3, ws.max_row + 2):
                key = key_fn(r) if r <= ws.max_row else None
                if key != last_key:
                    if r - start > 1:
                        ws.merge_cells(start_row=start, start_column=col,
                                        end_row=r-1, end_column=col)
                    start = r
                    last_key = key

        # Merge Time cells: merge vertically for consecutive identical (Date, Time) pairs
        merge_vertical(2, lambda r: (ws.cell(r,1).value, ws.cell(r,2).value))
        # Merge Date cells: merge vertically across all rows for that day
        merge_vertical(1, lambda r: ws.cell(r, 1).value)



        # ------------ DEFINE fills ------------
        yellow = PatternFill('solid', fgColor='FFFF54')  # bright yellow for Fixed modules
        red = PatternFill('solid', fgColor='EA3323')     # red-orange for Core modules
        blue = PatternFill('solid', fgColor='E0EAF6')    # light blue for alternating rows
        green = PatternFill('solid', fgColor='CBE9B8')   # light green for alternating rows

        # ------------ APPLY alternating row fills BY DAY ------------
        for excel_row, (d_idx, s_idx) in enumerate(row_meta, start=2):  # Excel rows start at 2 (after header)
            fill = blue if d_idx % 2 == 0 else green
            for col in range(1, 6):  # columns A(1) to E(5)
                ws.cell(row=excel_row, column=col).fill = fill

        # ------------ APPLY fixed/core exam coloring (overwrites cols 3-5) ------------
        for r in range(2, ws.max_row + 1):
            exam_name = ws.cell(r, 3).value
            fill = None
            if exam_name:
                # Check fixed modules (yellow)
                if any(exam_name.startswith(fm) for fm in Fixed_modules):
                    fill = yellow
                # Check core modules (red) overrides yellow
                if any(exam_name.startswith(cm) for cm in Core_modules):
                    fill = red
            if fill:
                for c in (3, 4, 5):  # Exam, Total No Students, Room columns
                    ws.cell(r, c).fill = fill
        #Centre text

        for row in range(2, ws.max_row + 1):
            for col in [1, 2]:
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(vertical='center')
        # ------------ SAVE workbook ------------
        wb.save(filename)
        


    st.write(f"Excel file '{filename}' created with merged cells, colors, and full schedule.")


def animation_html():
    return """
    <div class="wrapper">
      <div class="container" id="container">
        <div class="flange-top">
          <div class="flange-fill" id="flangeTopFill"></div>
        </div>
        <div class="flange-bottom">
          <div class="flange-fill" id="flangeBottomFill"></div>
        </div>
        <div class="i-body">
          <div class="fill" id="bodyFill"></div>
        </div>
      </div>
    </div>

<style>
  .wrapper {
    width: 267px;
    height: 267px;
    background-color: #98fb98;
    border-radius: 50%;
    display: flex;
    justify-content: center;
    align-items: center;
    margin: 40px auto;
    overflow: hidden;
  }

  .container {
    width: 267px;
    height: 267px;
    position: relative;
    transform-origin: center;
    transition: transform 1s ease-in-out;
  }

  .flange-top,
  .flange-bottom {
    width: 117px;
    height: 30px;
    background: white;
    position: absolute;
    left: 75px;
    overflow: hidden;
  }

  .flange-top {
    top: 40px;
  }

  .flange-bottom {
    bottom: 40px;
  }

  .flange-fill {
    position: absolute;
    bottom: 0;
    left: 0;
    width: 100%;
    height: 0%;
    background: blue;
    transition: height 1s ease-in-out;
  }

  .i-body {
    position: absolute;
    top: 70px;
    height: 127px;
    width: 31px;
    left: 118px;
    background: white;
    overflow: hidden;
    box-sizing: border-box;
  }

  .fill {
    position: absolute;
    bottom: 0;
    width: 100%;
    height: 0%;
    background: blue;
    transition: height 1s ease-in-out;
  }
</style>


<script>
  let angle = 0;
  const delay = (ms) => new Promise(resolve => setTimeout(resolve, ms));

  const container = document.getElementById("container");
  const bodyFill = document.getElementById("bodyFill");
  const flangeTopFill = document.getElementById("flangeTopFill");
  const flangeBottomFill = document.getElementById("flangeBottomFill");

  function setFillDirection(element, fromTop) {
    if (fromTop) {
      element.style.top = '0';
      element.style.bottom = '';
    } else {
      element.style.bottom = '0';
      element.style.top = '';
    }
  }

  async function fillSequence1(fromTop) {
    setFillDirection(flangeBottomFill, fromTop);
    flangeBottomFill.style.height = "100%";
    await delay(1000);

    setFillDirection(bodyFill, fromTop);
    bodyFill.style.height = "100%";
    await delay(1000);

    setFillDirection(flangeTopFill, fromTop);
    flangeTopFill.style.height = "100%";
    await delay(1000);
  }

  async function emptySequence1(fromTop) {
    setFillDirection(flangeTopFill, fromTop);
    flangeTopFill.style.height = "0%";
    await delay(1000);

    setFillDirection(bodyFill, fromTop);
    bodyFill.style.height = "0%";
    await delay(1000);

    setFillDirection(flangeBottomFill, fromTop);
    flangeBottomFill.style.height = "0%";
    await delay(1000);
  }
   async function fillSequence2(fromTop) {
    setFillDirection(flangeTopFill, fromTop);
    flangeTopFill.style.height = "100%";
    await delay(1000);



    setFillDirection(bodyFill, fromTop);
    bodyFill.style.height = "100%";
    await delay(1000);

    setFillDirection(flangeBottomFill, fromTop);
    flangeBottomFill.style.height = "100%";
    await delay(1000);
  }

  async function emptySequence2(fromTop) {
    setFillDirection(flangeBottomFill, fromTop);
    flangeBottomFill.style.height = "0%";
    await delay(1000);


    setFillDirection(bodyFill, fromTop);
    bodyFill.style.height = "0%";
    await delay(1000);

    setFillDirection(flangeTopFill, fromTop);
    flangeTopFill.style.height = "0%";
    await delay(1000);
  }

  async function animateCycle() {
    while (true) {
      // Fill bottom-to-top visually (normal orientation)
      await fillSequence1(false);
      await delay(500);

      angle += 180;
      container.style.transform = `rotate(${angle}deg)`;
      await delay(1000);

      // Empty bottom-to-top visually (but now rotated, so DOM-top is visual-bottom)
      await emptySequence2(true);
      await delay(500);

      // Fill bottom-to-top visually (still rotated, so fill from top in DOM)
      await fillSequence2(true);
      await delay(500);

      angle += 180;
      container.style.transform = `rotate(${angle}deg)`;
      await delay(1000);

      // Empty bottom-to-top visually (now upright, so DOM-bottom is visual-bottom)
      await emptySequence1(false);
      await delay(500);
    }
  }

  // Start animation after DOM ready
  setTimeout(() => {
    animateCycle();
  }, 100);
</script>
    """


# Main execution flow
if __name__ == "__main__":
    # Add a generate button
    if st.button("Generate Timetable"):
        # Process the files
        students_df, leaders_df, wb = process_files()
        if not all([student_file, module_file, dates_file]):
            st.error("Please upload all required files first.")
        else:
            try:
                animation_placeholder = st.empty()
                result_container = st.empty()
                processing_done = False
                error_msg = None

                def generate():
                    global processing_done, error_msg, students_df, leaders_df
                    try:
           

                        # Generate the timetable
                        timetable, days, exam_counts = create_timetable(
                            students_df, leaders_df, wb, max_exams_2days, max_exams_5days,
                        )
                        # Create the Excel file
                        generate_excel(timetable, days, exam_counts)
                    except Exception as e:
                        error_msg = str(e)
                    finally:
                        processing_done = True

                # Start background thread
                thread = threading.Thread(target=generate)
                thread.start()

                # Looping animation while waiting
                while not processing_done:
                    with animation_placeholder:
                        components.html(animation_html(), height=350)
                    time.sleep(2.1)

                animation_placeholder.empty()

                if error_msg:
                    st.error(f"An error occurred: {error_msg}")
                    logger.error(f"Error generating timetable: {error_msg}", exc_info=True)
                else:
                    # Success output
                    st.success("✅ Timetable generated successfully!")

                    with open("exam_schedule_merged0.xlsx", "rb") as file:
                        st.download_button(
                            label="Download Timetable",
                            data=file,
                            file_name="exam_schedule.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                    st.header("Generated Timetable")
                    df = pd.read_excel("exam_schedule_merged0.xlsx")
                    st.dataframe(df)

            except Exception as e:
                st.error(f"Unexpected error: {str(e)}")
