# %% [markdown]
# Going to use pandas to read the excel files

# %%
# %pip install pandas
# %pip install rapidfuzz
# %pip install collections
# %pip install openpyxl

# %%
import pandas as pd
from rapidfuzz import process, fuzz
from collections import defaultdict
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re
from dateutil.parser import parse

# %% [markdown]
# Read the files

# %%
students_df = pd.read_excel('/Users/edwardbrady/Library/CloudStorage/OneDrive-ImperialCollegeLondon/UROP-Exam timetabling/input data/student list DONOT SORT ONLY FILTER.xlsx',header=None)
leaders_df = pd.read_excel('/Users/edwardbrady/Library/CloudStorage/OneDrive-ImperialCollegeLondon/UROP-Exam timetabling/input data/module list 2024-25.xlsx',sheet_name=1,header=1)
# Load workbook and worksheet for useful dates
wb = load_workbook("/Users/edwardbrady/Library/CloudStorage/OneDrive-ImperialCollegeLondon/UROP-Exam timetabling/input data/2025-26 Useful Dates.xlsx")
ws = wb.active

# %% [markdown]
# Creating a list of exam names, from excel can see they're in Row 'A' which is 0 and start from column 'J' which is 9

# %%
exam_names = students_df.iloc[0, 9:].dropna().tolist()
print(exam_names)

# %% [markdown]
# To calculate the number of days for the exam period we can assume it will be 3 weeks, weeks 31-33 and therefore it will be 21 days starting on monday of week 31

# %%
num_days = 21

# %% [markdown]
# To create a dictionary of students and there exams we will flick through each row, which is a studnet, and for every module that has an x , A or B add it to their exams. At the end add PEN as everyone does PEN
# 

# %%
# Get the range of rows containing student data (from row 3 onward)
student_rows = students_df.iloc[2:, :]  # row index 3 and onward

# Initialize the dictionary
student_exams = {}

for _, row in student_rows.iterrows():
    cid = row[0]  # Column A = student CID
    exams_taken = []

    for col_idx, exam_name in enumerate(exam_names, start=9):  # Column J = index 9
        if str(row[col_idx]).strip().lower() == 'x':  # Check for 'x' (case-insensitive)
            exams_taken.append(exam_name)

    #Add in PEN exam
    exams_taken.append('MECH60015 / MECH70030 Professional Engineering Skills (ME3/AME) (ECE exam)')
    student_exams[cid] = exams_taken



print(student_exams)

# %% [markdown]
# To create a list of core modules I will read it off the excel from last year 

# %%
Core_modules = ["MECH70001 Nuclear Thermal Hydraulics","MECH60004/MECH70042 Introduction to Nuclear Energy A/B","MECH70002 Nuclear Reactor Physics","MECH70008 Mechanical Transmissions Technology","MECH70006 Metal Processing Technology","MECH70021Aircraft Engine Technology","MECH70003 Future Clean Transport Technology","MECH60015 / MECH70030 Professional Engineering Skills (ME3/AME) (ECE exam)"]

# %% [markdown]
# Creating a dictionary of fixed modules and their dates, in the format day, slot with day being the day from the first monday (including w/e) of exam season and slot being either morning or afternoon

# %%
Fixed_modules = {"BUSI60039 Business Strategy" :[1,1],"BUSI60046 Project Management":[2,1],"ME-ELEC70098 Optimisation":[3,0],"MECH70001 Nuclear Thermal Hydraulics":[3,0],"BUSI60040/BUSI60043 Corporate Finance Online/Finance & Financial Management":[3,1],"MECH60004/MECH70042 Introduction to Nuclear Energy A/B":[4,0],"ME-ELEC70022 Modelling and Control of Multi-body Mechanical Systems":[4,0],"MATE97022 Nuclear Materials 1":[4,0],"ME-MATE70029 Nuclear Fusion":[9,0],"MECH70002 Nuclear Reactor Physics":[10,0],"ME-ELEC70076 Sustainable Electrical Systems":[10,0],"ME ELEC70066 Applied Advanced Optimisation":[10,0],"MECH70020 Combustion, Safety and Fire Dynamics":[11,0],"BIOE70016 Human Neuromechanical Control and Learning":[11,0],"CENG60013 Nuclear Chemical Engineering":[11,0],"MECH70008 Mechanical Transmissions Technology":[13,1],"MECH70006 Metal Processing Technology":[13,1],"MECH70021Aircraft Engine Technology":[13,1],"MECH70003 Future Clean Transport Technology":[13,1],"MECH60015 / MECH70030 Professional Engineering Skills (ME3/AME) (ECE exam)":[14,1]}

# %% [markdown]
# Creat a dictionary of each module leader and there respective exams, this needs to use the names used in the students spreadsheet and as such we combine course code and name and find the closest match. Some are disregarded as they're not examinable courses.

# %%

# Extract official module names from row 0, columns J onwards (i.e., column 9 onward, 0-indexed)
standardized_names = students_df.iloc[0, 9:].dropna().astype(str).str.strip().tolist()


# Prepare module-leader dictionary
leader_courses = defaultdict(list)

# Loop through rows in the module list
for _, row in leaders_df.iterrows():
    leader = row['Module Leader (lecturer 1)']
    name = row['Module Name']
    code = row['Banner Code (New CR)']   # module leader

    # Skip if any required field is missing
    if pd.isna(code) or pd.isna(name) or pd.isna(leader):
        continue

    if leader == "n/a":
        continue

    # Combine code and name
    combined_name = f"{code} {name}"

    # Fuzzy match to standardized names
    best_match, score, _ = process.extractOne(
        combined_name, standardized_names, scorer=fuzz.token_sort_ratio
    )

    if score >= 70:
        if best_match not in leader_courses[leader]:
            leader_courses[leader].append(best_match)
        else:
            print(f"⚠️ Duplicate match skipped for '{combined_name}': '{best_match}' is already listed for {leader}.")
    else:
        print(f"⚠️ Low confidence match for '{combined_name}' (best: '{best_match}', score: {score}).")


# Convert to normal dict if desired
leader_courses = dict(leader_courses)

print("Module leaders and their courses:")
print(leader_courses)

# %% [markdown]
# Make a list of students with 25% extra time to ensure they dont have more than on exam a day

# %%
extra_time_students_25 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("15min/hour", "25% extra time"))].iloc[:, 0].tolist()

print(extra_time_students_25)

# %%
extra_time_students_50 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("30min/hour", "50% extra time"))].iloc[:, 0].tolist()

print(extra_time_students_50)

# %% [markdown]
# Need to read the bank holidays and forbidden time slots these will have the form [day,slot] the weekends, (days 5, 6 and 12, 13) can be assumed to not have exams and the no exam on morning of the last friday

# %%
no_exam_dates = [[5,0],[5,1],[6,0],[6,1],[12,0],[12,1],[13,0],[13,1],[20,0],]

# %% [markdown]
# To find the bank holidays we will use the useful dates spreadsheet as this has the start of the summer term and also the bank holidays, first we find the date of the start of summer term, this is made more difficult due to the fact that it is written in the form of day, month to day month year and we need day month year of start. Once this is found the first monday is assumed to be the following monday from this. Then the dates of the bank holidays are found and anything within 20 days is found and appended to the no exam dates

# %%


# Initialize
bank_holidays = []


# --- Step 1: Extract bank holidays from col F (names) and G (dates) ---
row = 5
while True:
    name = ws[f"F{row}"].value
    date_cell = ws[f"G{row}"].value
    if name is None or "Term Dates" in str(name):
        break
    if isinstance(date_cell, datetime):
        bank_holidays.append((str(name).strip(), date_cell.date()))
    row += 1

# --- Step 2: Find Summer Term start date from section below ---
summer_start = None
while row < ws.max_row:
    cell_value = ws[f"F{row}"].value
    if cell_value and "Summer Term" in str(cell_value):
        term_range = ws[f"F{row + 1}"].value
        if term_range:
            try:
                # Extract left side of range before "to", drop weekday (e.g., "Fri"), and append year
                start_part = term_range.split("to")[0].strip()
                start_str = re.sub(r"^\w+\s+", "", start_part)  # Removes "Fri", leaves "24 Apr"
                # Try extracting year from second part if present
                year_match = re.search(r"\b\d{4}\b", term_range)
                if year_match:
                    start_str += f" {year_match.group(0)}"
                else:
                    raise ValueError("Year not found in date range.")
                summer_start = parse(start_str, dayfirst=True).date()
            except Exception as e:
                raise ValueError(f"Could not parse Summer Term start: {term_range}") from e
        else:
            raise ValueError("Summer Term range cell is empty.")
        break
    row += 1

if not summer_start:
    raise ValueError("Summer Term start date not found.")

# --- Step 3: Find first Monday on or after summer_start ---
first_monday = summer_start
while first_monday.weekday() != 0:  # 0 = Monday
    first_monday += timedelta(days=1)

# --- Step 4: Find bank holidays within 3 weeks after first Monday ---
for name, bh_date in bank_holidays:
    delta = (bh_date - first_monday).days
    if 0 <= delta <= 20:
        print(f"{name} is {delta} days after first Monday ({bh_date})")
        no_exam_dates.append([delta, 0])
        no_exam_dates.append([delta, 1])

# --- Final Output ---
print("No exam dates:", no_exam_dates)



