

#Imports for data reading
import pandas as pd
from rapidfuzz import process, fuzz
from collections import defaultdict
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re
from dateutil.parser import parse
import os


#Imports for model
from ortools.sat.python import cp_model

#Imports for the excel file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
#Imports for checking
from collections import defaultdict

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

primaryColor = "#00ff7f"
backgroundColor = "#ffffff"
secondaryBackgroundColor = "#f5f5f5"
textColor = "#0000cd"
font = "sans serif"


# In[104]:
class ExamTimetableUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Exam Timetabling System")
        self.root.configure(bg=backgroundColor)

        self.files = {"student": None, "module": None, "dates": None}
        self.output_folder = None
        self.test_files = []

        # --- File Upload Frame ---
        file_frame = tk.Frame(root, bg=secondaryBackgroundColor, padx=10, pady=10)
        file_frame.grid(row=0, column=0, columnspan=3, sticky="ew", padx=10, pady=10)

        tk.Label(file_frame, text="Upload Required Files", font=(font, 16, "bold"), bg=secondaryBackgroundColor, fg=textColor).grid(row=0, column=0, columnspan=3, pady=5)

        tk.Button(file_frame, text="Upload Student List", bg=primaryColor, fg=textColor, font=(font, 12, "bold"), command=self.upload_student).grid(row=1, column=0, padx=5, pady=5)
        tk.Button(file_frame, text="Upload Module List", bg=primaryColor, fg=textColor, font=(font, 12, "bold"), command=self.upload_module).grid(row=1, column=1, padx=5, pady=5)
        tk.Button(file_frame, text="Upload Useful Dates", bg=primaryColor, fg=textColor, font=(font, 12, "bold"), command=self.upload_dates).grid(row=1, column=2, padx=5, pady=5)

        # --- Parameters Frame ---
        param_frame = tk.Frame(root, bg=secondaryBackgroundColor, padx=10, pady=10)
        param_frame.grid(row=1, column=0, columnspan=3, sticky="ew", padx=10, pady=10)

        tk.Label(param_frame, text="Timetabling Parameters", font=(font, 16, "bold"), bg=secondaryBackgroundColor, fg=textColor).grid(row=0, column=0, columnspan=2, pady=5)

        tk.Label(param_frame, text="Number of Days for Exam Period:", font=(font, 12), bg=secondaryBackgroundColor, fg=textColor).grid(row=1, column=0, sticky="w", pady=5)
        self.num_days = tk.Spinbox(param_frame, from_=1, to=30, font=(font, 12), value=21)
        self.num_days.grid(row=1, column=1, pady=5)


        tk.Label(param_frame, text="Max Exams in 2-Day Window:", font=(font, 12), bg=secondaryBackgroundColor, fg=textColor).grid(row=2, column=0, sticky="w", pady=5)
        self.max_exams_2days = tk.Spinbox(param_frame, from_=1, to=5, font=(font, 12), value=3)
        self.max_exams_2days.grid(row=2, column=1, pady=5)

        tk.Label(param_frame, text="Max Exams in 5-Day Window:", font=(font, 12), bg=secondaryBackgroundColor, fg=textColor).grid(row=3, column=0, sticky="w", pady=5)
        self.max_exams_5days = tk.Spinbox(param_frame, from_=1, to=10, font=(font, 12), value=4)
        self.max_exams_5days.grid(row=3, column=1, pady=5)
        
        # --- Soft Constraints Frame ---
        soft_frame = tk.Frame(root, bg=secondaryBackgroundColor, padx=10, pady=10)
        soft_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=10, pady=10)

        tk.Label(soft_frame, text="Soft Constraint Weights", font=(font, 14, "bold"), bg=secondaryBackgroundColor, fg=textColor).grid(row=0, column=0, columnspan=2, pady=5)

        # Sliders
        self.spread_penalty = tk.IntVar(value=5)
        tk.Label(soft_frame, text="Module leaders exams spread out:", font=(font, 12), bg=secondaryBackgroundColor, fg=textColor).grid(row=1, column=0, sticky="w", pady=5)
        tk.Scale(soft_frame, from_=0, to=10, orient="horizontal", variable=self.spread_penalty, bg=secondaryBackgroundColor, troughcolor=primaryColor).grid(row=1, column=1, pady=5)

        self.room_penalty = tk.IntVar(value=5)
        tk.Label(soft_frame, text="More than 2 rooms per exam:", font=(font, 12), bg=secondaryBackgroundColor, fg=textColor).grid(row=2, column=0, sticky="w", pady=5)
        tk.Scale(soft_frame, from_=0, to=10, orient="horizontal", variable=self.room_penalty, bg=secondaryBackgroundColor, troughcolor=primaryColor).grid(row=2, column=1, pady=5)

        self.extra_time_penalty = tk.IntVar(value=5)
        tk.Label(soft_frame, text="25% Extra Time students multiple exams:", font=(font, 12), bg=secondaryBackgroundColor, fg=textColor).grid(row=3, column=0, sticky="w", pady=5)
        tk.Scale(soft_frame, from_=0, to=10, orient="horizontal", variable=self.extra_time_penalty, bg=secondaryBackgroundColor, troughcolor=primaryColor).grid(row=3, column=1, pady=5)

        self.soft_day_penalty = tk.IntVar(value=5)
        tk.Label(soft_frame, text="No exams on certain days:", font=(font, 12), bg=secondaryBackgroundColor, fg=textColor).grid(row=4, column=0, sticky="w", pady=5)
        tk.Scale(soft_frame, from_=0, to=10, orient="horizontal", variable=self.soft_day_penalty, bg=secondaryBackgroundColor, troughcolor=primaryColor).grid(row=4, column=1, pady=5)

        # --- Generate Button ---
        tk.Button(root, text="Generate Timetable", bg=primaryColor, fg=textColor, font=(font, 12, "bold"), command=self.generatetimetable).grid(
            row=3, column=0, columnspan=3, pady=20
        )

        


    # --- File upload methods ---
    def upload_student(self):
        self.files["student"] = filedialog.askopenfilename(title="Select Student List", filetypes=[("Excel files", "*.xlsx")])
        print("Student List:", self.files["student"])

    def upload_module(self):
        self.files["module"] = filedialog.askopenfilename(title="Select Module List", filetypes=[("Excel files", "*.xlsx")])
        print("Module List:", self.files["module"])

    def upload_dates(self):
        self.files["dates"] = filedialog.askopenfilename(title="Select Useful Dates", filetypes=[("Excel files", "*.xlsx")])
        print("Useful Dates:", self.files["dates"])

    def select_output_folder(self):
        self.output_folder = filedialog.askdirectory(title="Select Output Folder")

    def select_test_files(self):
        self.test_files = filedialog.askopenfilenames(title="Select Test Files", filetypes=[("Excel files", "*.xlsx")])




    def generatetimetable(self):


        # Opening all the excel spreadsheets
        students_df = pd.read_excel(app.files["student"], header=None)
        leaders_df = pd.read_excel(app.files["module"], sheet_name=1, header=1)
        wb = load_workbook(app.files["dates"])
        ws = wb.active
        num_days = int(app.num_days.get())
        max_exams_2days = int(app.max_exams_2days.get())
        max_exams_5days = int(app.max_exams_5days.get())
        spread_penalty = app.spread_penalty.get()
        room_penalty = app.room_penalty.get()
        extra_time_penalty = app.extra_time_penalty.get()
        soft_day_penalty = app.soft_day_penalty.get


        # Creating a list of exam names, from excel can see they're in Row 'A' which is 0 and start from column 'J' which is 9

        exams = students_df.iloc[0, 9:].dropna().tolist()



        # To calculate the number of days for the exam period we can assume it will be 3 weeks, weeks 31-33 and therefore it will be 21 days starting on monday of week 31
        num_days = 20

        # To create a dictionary of students and there exams we will flick through each row, which is a studnet, and for every module that has an x , A or B add it to their exams. At the end add PEN as everyone does PEN
        # Get the range of rows containing student data (from row 3 onward)
        student_rows = students_df.iloc[2:, :]  # row index 3 and onward



        # Create a boolean mask for column D where value is not "#N/A" (after stripping whitespace)
        valid_aea_mask = (
            student_rows.iloc[:, 3].notna() &
            (student_rows.iloc[:, 3].astype(str).str.strip() != "#N/A")
        )

        # Get the CIDs (from column 0) where AEA condition is met
        AEA = student_rows.loc[valid_aea_mask, student_rows.columns[0]].tolist()


        # Initialize the dictionary
        student_exams = {}

        for _, row in student_rows.iterrows():
            cid = row[0]  # Column A = student CID
            exams_taken = []

            for col_idx, exam_name in enumerate(exams, start=9):  # Column J = index 9
                if str(row[col_idx]).strip().lower() == 'x' or str(row[col_idx]).strip().lower() == 'a'  or str(row[col_idx]).strip().lower() == 'b' :  # Check for 'x' or 'a' or 'b' to indicate they take this course (case-insensitive)
                    exams_taken.append(exam_name)

            student_exams[cid] = exams_taken






        # In[106]:


        # To create a list of core modules I will read it off the excel from last year

        # %%
        Core_modules = ["MECH70001 Nuclear Thermal Hydraulics","MECH60004/MECH70042 Introduction to Nuclear Energy A/B","MECH70002 Nuclear Reactor Physics","MECH70008 Mechanical Transmissions Technology","MECH70006 Metal Processing Technology","MECH70021Aircraft Engine Technology","MECH70003 Future Clean Transport Technology","MECH60015/70030 PEN3/AME"]


        # Creating a dictionary of fixed modules and their dates, in the format day, slot with day being the day from the first monday (including w/e) of exam season and slot being either morning or afternoon
        Fixed_modules = {"BUSI60039 Business Strategy" :[1,1],"BUSI60046 Project Management":[2,1],"ME-ELEC70098 Optimisation":[3,0],"MECH70001 Nuclear Thermal Hydraulics":[3,0],"BUSI60040/BUSI60043 Corporate Finance Online/Finance & Financial Management":[3,1],"MECH60004/MECH70042 Introduction to Nuclear Energy A/B":[4,0],"ME-ELEC70022 Modelling and Control of Multi-body Mechanical Systems":[4,0],"MATE97022 Nuclear Materials 1":[4,0],"ME-MATE70029 Nuclear Fusion":[9,0],"MECH70002 Nuclear Reactor Physics":[10,0],"ME-ELEC70076 Sustainable Electrical Systems":[10,0],"ME ELEC70066 Applied Advanced Optimisation":[10,0],"MECH70020 Combustion, Safety and Fire Dynamics":[11,0],"BIOE70016 Human Neuromechanical Control and Learning":[11,0],"CENG60013 Nuclear Chemical Engineering":[11,0],"MECH70008 Mechanical Transmissions Technology":[17,1],"MECH70006 Metal Processing Technology":[17,1],"MECH70021Aircraft Engine Technology":[17,1],"MECH70003 Future Clean Transport Technology":[17,1],"MECH60015/70030 PEN3/AME":[18,1]}



        # Creat a dictionary of each module leader and there respective exams, this needs to use the names used in the students spreadsheet and as such we combine course code and name and find the closest match. Some are disregarded as they're not examinable courses.


        # Extract official module names from row 0, columns J onwards (i.e., column 9 onward, 0-indexed)
        standardized_names = exams


        # Prepare module-leader dictionary
        leader_courses = defaultdict(list)
        exam_types = dict()
        # Loop through rows in the module list
        for _, row in leaders_df.iterrows():

            leaders = []
            if pd.notna(row['Module Leader (lecturer 1)']):
                leaders.append(row['Module Leader (lecturer 1)'])

            if pd.notna(row['(UGO Internal) 2nd Exam Marker']):
                leaders.append(row['(UGO Internal) 2nd Exam Marker'])

            name = row['Module Name']
            code = row['Banner Code (New CR)']   # module leader

            # Skip if any required field is missing
            if pd.isna(code) or pd.isna(name) :
                continue

            if len(leaders) == 0 :
                continue

            # Combine code and name
            combined_name = f"{code} {name}"

            # Fuzzy match to standardized names
            best_match, score, _ = process.extractOne(
                combined_name, standardized_names, scorer=fuzz.token_sort_ratio
            )

            if score >= 70: 
                exam_types[best_match] = row['(UGO Internal) Exam Style'] if pd.notna(row['(UGO Internal) Exam Style']) else None
                for leader in leaders:

                    if best_match not in leader_courses[leader]:
                        leader_courses[leader].append(best_match)
                    else:
                        a=1
                        # print(f"⚠️ Duplicate match skipped for '{combined_name}': '{best_match}' is already listed for {leader}.")
            else:
                a=1
                # print(f"⚠️ Low confidence match for '{combined_name}' (best: '{best_match}', score: {score}).")


        # Convert to normal dict if desired
        leader_courses = dict(leader_courses)

        for exam in exams:
            if exam not in exam_types:
                exam_types[exam] = "Standard"




        # Make a list of students with 25% extra time to ensure they dont have more than on exam a day

        extra_time_students_25 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("15min/hour", "25% extra time"))].iloc[:, 0].tolist()



        # Make a list of students with 50% extra time to ensure they dont have more than on exam a day
        extra_time_students_50 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("30min/hour", "50% extra time"))].iloc[:, 0].tolist()




        # Need to read the bank holidays and forbidden time slots these will have the form [day,slot] the weekends, (days 5, 6 and 12, 13) can be assumed to not have exams and the no exam on morning of the last friday

        no_exam_dates = [[5,0],[5,1],[6,0],[6,1],[12,0],[12,1],[13,0],[13,1],[18,0],[19,0],[19,1],[20,0],[20,1],]

        no_exam_dates_soft = [
            [15,0],# Week 3 tuesday morning
            [16,0], #Week 3 Wednesday morning
        ]

        # To find the bank holidays we will use the useful dates spreadsheet as this has the start of the summer term and also the bank holidays, first we find the date of the start of summer term, this is made more difficult due to the fact that it is written in the form of day, month to day month year and we need day month year of start. Once this is found the first monday is assumed to be the following monday from this. Then the dates of the bank holidays are found and anything within 20 days is found and appended to the no exam dates

        # Initialize
        bank_holidays = []


        # --- Step 1: Extract bank holidays from col F (names) and G (dates) ---
        row = 5
        while True:
            name = ws[f"F{row}"].value
            date_cell = ws[f"G{row}"].value
            if name is None or "Term Dates" in str(name):
                break
            if isinstance(date_cell, datetime):
                bank_holidays.append((str(name).strip(), date_cell.date()))
            row += 1

        # --- Step 2: Find Summer Term start date from section below ---
        summer_start = None
        while row < ws.max_row:
            cell_value = ws[f"F{row}"].value
            if cell_value and "Summer Term" in str(cell_value):
                term_range = ws[f"F{row + 1}"].value
                if term_range:
                    try:
                        # Extract left side of range before "to", drop weekday (e.g., "Fri"), and append year
                        start_part = term_range.split("to")[0].strip()
                        start_str = re.sub(r"^\w+\s+", "", start_part)  # Removes "Fri", leaves "24 Apr"
                        # Try extracting year from second part if present
                        year_match = re.search(r"\b\d{4}\b", term_range)
                        if year_match:
                            start_str += f" {year_match.group(0)}"
                        else:
                            raise ValueError("Year not found in date range.")
                        summer_start = parse(start_str, dayfirst=True).date()
                    except Exception as e:
                        raise ValueError(f"Could not parse Summer Term start: {term_range}") from e
                else:
                    raise ValueError("Summer Term range cell is empty.")
                break
            row += 1

        if not summer_start:
            raise ValueError("Summer Term start date not found.")

        # --- Step 3: Find first Monday on or after summer_start ---
        first_monday = summer_start
        while first_monday.weekday() != 0:  # 0 = Monday
            first_monday += timedelta(days=1)

        # --- Step 4: Find bank holidays within 3 weeks after first Monday ---
        for name, bh_date in bank_holidays:
            delta = (bh_date - first_monday).days
            if 0 <= delta <= 20:
                no_exam_dates.append([delta, 0])
                no_exam_dates.append([delta, 1])

        # --- Final Output ---
        print("No exam dates:", no_exam_dates)

        def ordinal(n):
            # Returns ordinal string for an integer n, e.g. 1 -> 1st, 2 -> 2nd
            if 11 <= (n % 100) <= 13:
                return f"{n}th"
            else:
                return f"{n}{['th','st','nd','rd','th','th','th','th','th','th'][n % 10]}"

        days = []
        for i in range(21):
            date = first_monday + timedelta(days=i)
            day_str = date.strftime("%A ") + ordinal(date.day) + date.strftime(" %B")
            days.append(day_str)

        #Dictionary of rooms uses and capacity
        # Room dictionary with capacities and functions
        rooms = {
            'CAGB 203': [["Computer", "SEQ"], 65],
            'CAGB 309': [["SEQ"], 54],
            'CAGB 649-652': [["SEQ"], 75],
            'CAGB 747-748': [["AEA"], 36],
            'CAGB 749-752': [["SEQ"], 75],
            'CAGB 761': [["Computer", "AEA"], 25],
            'CAGB 762': [["Computer", "AEA"], 25],
            'CAGB 765': [["AEA","Computer"], 10],
            'CAGB 527': [["AEA"], 2],
            'NON ME N/A':[["SEQ","AEA"],1000] #For business modules
        }

        #Total student count for AEA and non AEA
        exam_counts = defaultdict(lambda: [0, 0])

        for cid, exams_taken in student_exams.items():
            if cid in AEA:
                for exam in exams_taken:
                    exam_counts[exam][0] += 1
            else:
                for exam in exams_taken:
                    exam_counts[exam][1] += 1




        # After we have data we need to initialize the model

        # In[107]:


        slots = [0,1]

        num_slots = len(slots)


        # In[108]:


        # Initailize the model
        model = cp_model.CpModel()


        # Variables: exam_day and exam_slot
        exam_day = {}
        exam_slot = {}
        for exam in exams:
            exam_day[exam] = model.NewIntVar(0, num_days - 1, f'{exam}_day')
            exam_slot[exam] = model.NewIntVar(0, num_slots - 1, f'{exam}_slot')

        # Variable to put exams in rooms
        exam_room = {}

        for exam in exams:
            for room in rooms:
                exam_room[(exam, room)] = model.NewBoolVar(f'{exam}_in_{room.replace(" ", "_")}')


        # After model is initialized we add the constraints for the exams

        # In[109]:


        # 0. Students can't have exams at the same tiem
        for student, exs in student_exams.items():
        #Loops through students
            for i in range(len(exs)):
                for j in range(i + 1, len(exs)):
                    exam1 = exs[i]
                    exam2 = exs[j]
                    #Boolean variables for day and slot matches
                    same_day = model.NewBoolVar(f'{exam1}_same_day{exam2}')
                    same_slot = model.NewBoolVar(f'{exam1}_same_slot{exam2}')
                    
                    model.Add(exam_day[exam1] == exam_day[exam2]).OnlyEnforceIf(same_day)
                    model.Add(exam_day[exam1] != exam_day[exam2]).OnlyEnforceIf(same_day.Not())

                    model.Add(exam_slot[exam1] == exam_slot[exam2]).OnlyEnforceIf(same_slot)
                    model.Add(exam_slot[exam1] != exam_slot[exam2]).OnlyEnforceIf(same_slot.Not())

                    

                    model.AddBoolOr([same_day.Not(), same_slot.Not()])

        # 1. Core modules can not have multiple exams on that day
        for student, exs in student_exams.items():# Loop students
                # Create a list of their core exams
            core_mods = [exam for exam in exs if exam in Core_modules]
            other_mods = [exam for exam in exs if exam not in Core_modules]
                # Loop through the core exams
            for exam in core_mods:
                for other in other_mods:
                        # Add constraint
                    model.Add(exam_day[exam] != exam_day[other])


        # 2. Fixed modules day and slot assignment
        for exam, (day_fixed, slot_fixed) in Fixed_modules.items():
            model.Add(exam_day[exam] == day_fixed)
            model.Add(exam_slot[exam] == slot_fixed)




        # 3. Forbidden exam day-slot assignments

        for exam in exams:
            for day, slot in no_exam_dates:
                model.AddForbiddenAssignments([exam_day[exam], exam_slot[exam]], [(day, slot)])



        # 4. Max 3 exams in any 2-day window per student
        for student, ex in student_exams.items():
            for d in range(num_days - 1):
                exams_in_2_days = []
                for exam in ex:
                    is_on_d = model.NewBoolVar(f'{student}_{exam}_on_day_{d}')
                    is_on_d1 = model.NewBoolVar(f'{student}_{exam}_on_day_{d+1}')
                    is_on_either = model.NewBoolVar(f'{student}_{exam}_on_day_{d}_or_{d+1}')

                    model.Add(exam_day[exam] == d).OnlyEnforceIf(is_on_d)
                    model.Add(exam_day[exam] != d).OnlyEnforceIf(is_on_d.Not())

                    model.Add(exam_day[exam] == d + 1).OnlyEnforceIf(is_on_d1)
                    model.Add(exam_day[exam] != d + 1).OnlyEnforceIf(is_on_d1.Not())

                    model.AddBoolOr([is_on_d, is_on_d1]).OnlyEnforceIf(is_on_either)
                    model.AddBoolAnd([is_on_d.Not(), is_on_d1.Not()]).OnlyEnforceIf(is_on_either.Not())

                    exams_in_2_days.append(is_on_either)

                model.Add(sum(exams_in_2_days) <= 3)





        # 5. Max 4 exams in any 5-day sliding window per student
        for student, exs in student_exams.items():
            for start_day in range(num_days - 4):
                exams_in_window = []
                for exam in exs:
                    in_window = model.NewBoolVar(f'{student}_{exam}_in_day_{start_day}_to_{start_day + 4}')

                    model.AddLinearConstraint(exam_day[exam], start_day, start_day + 4).OnlyEnforceIf(in_window)

                    before_window = model.NewBoolVar(f'{student}_{exam}_before_{start_day}')
                    after_window = model.NewBoolVar(f'{student}_{exam}_after_{start_day + 4}')

                    model.Add(exam_day[exam] < start_day).OnlyEnforceIf(before_window)
                    model.Add(exam_day[exam] >= start_day).OnlyEnforceIf(before_window.Not())

                    model.Add(exam_day[exam] > start_day + 4).OnlyEnforceIf(after_window)
                    model.Add(exam_day[exam] <= start_day + 4).OnlyEnforceIf(after_window.Not())

                    model.AddBoolOr([before_window, after_window]).OnlyEnforceIf(in_window.Not())

                    exams_in_window.append(in_window)

                model.Add(sum(exams_in_window) <= 4)




        # 6. At most 1 exam in week 3 (days 13 to 20) per module leader
        for leader, leader_exams in leader_courses.items():
            exams_in_week3 = []
            for exam in leader_exams:
                in_week3 = model.NewBoolVar(f'{exam}_in_week3')

                model.AddLinearConstraint(exam_day[exam], 13, 20).OnlyEnforceIf(in_week3)

                before_week3 = model.NewBoolVar(f'{exam}_before_week3')
                after_week3 = model.NewBoolVar(f'{exam}_after_week3')

                model.Add(exam_day[exam] < 13).OnlyEnforceIf(before_week3)
                model.Add(exam_day[exam] >= 13).OnlyEnforceIf(before_week3.Not())

                model.Add(exam_day[exam] > 20).OnlyEnforceIf(after_week3)
                model.Add(exam_day[exam] <= 20).OnlyEnforceIf(after_week3.Not())

                model.AddBoolOr([before_week3, after_week3]).OnlyEnforceIf(in_week3.Not())

                exams_in_week3.append(in_week3)

            model.Add(sum(exams_in_week3) <= 1)




        # 7. Extra time 50% students: max 1 exam per day
        for student in extra_time_students_50:
            for day in range(num_days):
                exams_on_day = []
                for exam in student_exams[student]:
                    is_on_day = model.NewBoolVar(f'{student}_{exam}_on_day_{day}')
                    model.Add(exam_day[exam] == day).OnlyEnforceIf(is_on_day)
                    model.Add(exam_day[exam] != day).OnlyEnforceIf(is_on_day.Not())
                    exams_on_day.append(is_on_day)
                model.Add(sum(exams_on_day) <= 1)

        #Soft constraint that extra time students with<= 25% should only have one a day
        extra_time_25_penalties= []

        for student in extra_time_students_25:
            for day in range(num_days):
                exams_on_day = []

                for exam in student_exams[student]:
                    is_on_day = model.NewBoolVar(f'{student}_{exam}_on_day_{day}')
                    model.Add(exam_day[exam] == day).OnlyEnforceIf(is_on_day)
                    model.Add(exam_day[exam] != day).OnlyEnforceIf(is_on_day.Not())
                    exams_on_day.append(is_on_day)

                # Total number of exams student has that day
                num_exams = model.NewIntVar(0, len(exams_on_day), f'{student}_num_exams_day_{day}')
                model.Add(num_exams == sum(exams_on_day))

                # Create reified bool for "more than one exam"
                has_multiple_exams = model.NewBoolVar(f'{student}_more_than_one_exam_day_{day}')
                model.Add(num_exams >= 2).OnlyEnforceIf(has_multiple_exams)
                model.Add(num_exams < 2).OnlyEnforceIf(has_multiple_exams.Not())

                # Link this to a penalty variable (can just use the bool directly or an IntVar)
                penalty = model.NewIntVar(0, 1, f'{student}_penalty_day_{day}')
                model.Add(penalty == 1).OnlyEnforceIf(has_multiple_exams)
                model.Add(penalty == 0).OnlyEnforceIf(has_multiple_exams.Not())

                extra_time_25_penalties.append(penalty)

        #Soft constraint that course leaders modules should be spread out
        spread_penalties =[]

        for leader in leader_courses:#1 Loop through module leaders
            mods = leader_courses[leader]

            for i in range(len(mods)):#2 loop through modules
                for j in range(i+1, len(mods)):
                    m1 = mods[i]
                    m2 = mods[j]

                    #3 Calculate absolute day difference
                    diff = model.NewIntVar(-21, 21, f'{m1}_{m2}_diff')
                    abs_diff = model.NewIntVar(0, 21, f'{m1}_{m2}_abs_diff')
                                #4 Add difference to model
                    model.Add(diff == exam_day[m1] - exam_day[m2])
                    model.AddAbsEquality(abs_diff, diff)

                    #5 Create penalty variable
                    close_penalty = model.NewIntVar(0, 5, f'{m1}_{m2}_penalty')

                    #6 Create Boolean conditions
                    is_gap_3 = model.NewBoolVar(f'{m1}_{m2}_gap3')
                    is_gap_2 = model.NewBoolVar(f'{m1}_{m2}_gap2')
                    is_gap_1 = model.NewBoolVar(f'{m1}_{m2}_gap1')
                    is_gap_0 = model.NewBoolVar(f'{m1}_{m2}_gap0')
                                #7 Set the true condition
                    model.Add(abs_diff == 3).OnlyEnforceIf(is_gap_3)
                    model.Add(abs_diff != 3).OnlyEnforceIf(is_gap_3.Not())

                    model.Add(abs_diff == 2).OnlyEnforceIf(is_gap_2)
                    model.Add(abs_diff != 2).OnlyEnforceIf(is_gap_2.Not())

                    model.Add(abs_diff == 1).OnlyEnforceIf(is_gap_1)
                    model.Add(abs_diff != 1).OnlyEnforceIf(is_gap_1.Not())

                    model.Add(abs_diff == 0).OnlyEnforceIf(is_gap_0)
                    model.Add(abs_diff != 0).OnlyEnforceIf(is_gap_0.Not())

                    #8 Assign penalty values based gap
                    model.Add(close_penalty == 1).OnlyEnforceIf(is_gap_3)
                    model.Add(close_penalty == 3).OnlyEnforceIf(is_gap_2)
                    model.Add(close_penalty == 4).OnlyEnforceIf(is_gap_1)
                    model.Add(close_penalty == 5).OnlyEnforceIf(is_gap_0)

                    #9 no penalty if gap ≥ 4 and not equal to 0–3
                    model.Add(close_penalty == 0).OnlyEnforceIf(
                        is_gap_3.Not(), is_gap_2.Not(), is_gap_1.Not(), is_gap_0.Not()
                    )
                                #10 Add penalty to total penalties
                    spread_penalties.append(close_penalty)

        #Soft constraint to ensure no exams on some days
                    
        #1 Initiated list of penalties
        soft_day_penalties = []

        for exam in exams:#2 Loop exams
            for day, slot in no_exam_dates_soft:
                is_on_soft_day = model.NewBoolVar(f'{exam}_on_soft_day_{day}_{slot}')

                #3 Boolean variables for day and slot matches
                day_match = model.NewBoolVar(f'{exam}_day_eq_{day}')
                slot_match = model.NewBoolVar(f'{exam}_slot_eq_{slot}')
                
                model.Add(exam_day[exam] == day).OnlyEnforceIf(day_match)
                model.Add(exam_day[exam] != day).OnlyEnforceIf(day_match.Not())

                model.Add(exam_slot[exam] == slot).OnlyEnforceIf(slot_match)
                model.Add(exam_slot[exam] != slot).OnlyEnforceIf(slot_match.Not())

                #4 is_on_soft_day = day_match AND slot_match
                model.AddBoolAnd([day_match, slot_match]).OnlyEnforceIf(is_on_soft_day)
                model.AddBoolOr([day_match.Not(), slot_match.Not()]).OnlyEnforceIf(is_on_soft_day.Not())

                #5 Penalty: 10 if scheduled on soft day
                penalty = model.NewIntVar(0, 10, f'{exam}_penalty_soft_day_{day}_{slot}')
                model.Add(penalty == 10).OnlyEnforceIf(is_on_soft_day)
                model.Add(penalty == 0).OnlyEnforceIf(is_on_soft_day.Not())

                soft_day_penalties.append(penalty)

        #Minimize the amount of exams per slot 
        soft_slot_penalties = []

        for day in range(len(days)):  #1 First two weeks only
            for slot in slots:  
                exams_in_slot = []
                    
                        # 2 Make a list of all exams in a slot
                for exam in exams:
                    is_scheduled_day = model.NewBoolVar(f'{exam}_is_on_day{day}')
                    is_scheduled_slot = model.NewBoolVar(f'{exam}_is_on_slot{slot}')

                    model.Add(exam_day[exam] == day).OnlyEnforceIf(is_scheduled_day)
                    model.Add(exam_day[exam] != day).OnlyEnforceIf(is_scheduled_day.Not())

                    model.Add(exam_slot[exam] == slot).OnlyEnforceIf(is_scheduled_slot)
                    model.Add(exam_slot[exam] != slot).OnlyEnforceIf(is_scheduled_slot.Not())
                    
                    is_scheduled_here = model.NewBoolVar(f'{exam}_on_day{day}_slot{slot}')
                    model.AddBoolAnd([is_scheduled_day, is_scheduled_slot]).OnlyEnforceIf(is_scheduled_here)
                    model.AddBoolOr([is_scheduled_day.Not(), is_scheduled_slot.Not()]).OnlyEnforceIf(is_scheduled_here.Not())
                    exams_in_slot.append(is_scheduled_here)

                # 3 Count number of exams scheduled in this (day, slot)
                num_exams_here = model.NewIntVar(0, len(exams), f'count_day{day}_slot{slot}')
                model.Add(num_exams_here == sum(exams_in_slot))

                # 4 Calculate penalties
                is_three = model.NewBoolVar(f'is_three_day{day}_slot{slot}')
                is_four_or_more = model.NewBoolVar(f'is_four_plus_day{day}_slot{slot}')

                model.Add(num_exams_here == 3).OnlyEnforceIf(is_three)
                model.Add(num_exams_here != 3).OnlyEnforceIf(is_three.Not())

                model.Add(num_exams_here >= 4).OnlyEnforceIf(is_four_or_more)
                model.Add(num_exams_here < 4).OnlyEnforceIf(is_four_or_more.Not())

                #5 Apply penalties
                penalty_three = model.NewIntVar(0, 5, f'penalty_three_day{day}_slot{slot}')
                penalty_four = model.NewIntVar(0, 100, f'penalty_four_day{day}_slot{slot}')

                model.Add(penalty_three == 5).OnlyEnforceIf(is_three)
                model.Add(penalty_three == 0).OnlyEnforceIf(is_three.Not())

                model.Add(penalty_four == 100).OnlyEnforceIf(is_four_or_more)
                model.Add(penalty_four == 0).OnlyEnforceIf(is_four_or_more.Not())

                soft_slot_penalties.append(penalty_three)
                soft_slot_penalties.append(penalty_four)



        # Add constraints for the rooms

        # In[110]:


        for exam in exams:
            if exam in Fixed_modules and exam not in Core_modules:
                model.Add(exam_room[(exam, 'NON ME N/A')] ==1)  # Assign to N/A room if fixed module
            else:
                model.Add(exam_room[(exam, 'NON ME N/A')] == 0)  # Do not assign to N/A room if not fixed module


        for exam in exams: # Loop through
                    # Calculate capacity's for each room
            AEA_capacity = sum(
                rooms[room][1] * exam_room[(exam, room)]
                for room in rooms if "AEA" in rooms[room][0]
            )

            SEQ_capacity = sum(
                rooms[room][1] * exam_room[(exam, room)]
                for room in rooms if "SEQ" in rooms[room][0]
            )


            AEA_students = exam_counts[exam][0]
            SEQ_students = exam_counts[exam][1]

            # Add Constraint
            model.Add(AEA_capacity >= AEA_students)
            model.Add(SEQ_capacity >= SEQ_students)


        #Ensure only one exam is scheduled in each room at a time
        for d in range(num_days):
            for s in range(num_slots):
                for room in rooms:
                    if room == 'NON ME N/A':
                        continue  # Skip N/A room for this constraint  
                    else:
                        exams_in_room_time = []
                        for exam in exams:
                            # Only consider exams that *can* be scheduled in this day and slot
                            # Using model variables to express constraints:

                            # Create bool var: exam_at_time = (exam_day == d) AND (exam_slot == s)
                            exam_at_day = model.NewBoolVar(f'{exam}_on_day_{d}')
                            model.Add(exam_day[exam] == d).OnlyEnforceIf(exam_at_day)
                            model.Add(exam_day[exam] != d).OnlyEnforceIf(exam_at_day.Not())

                            exam_at_slot = model.NewBoolVar(f'{exam}_on_slot_{s}')
                            model.Add(exam_slot[exam] == s).OnlyEnforceIf(exam_at_slot)
                            model.Add(exam_slot[exam] != s).OnlyEnforceIf(exam_at_slot.Not())

                            exam_at_time = model.NewBoolVar(f'{exam}_on_{d}_{s}')
                            model.AddBoolAnd([exam_at_day, exam_at_slot]).OnlyEnforceIf(exam_at_time)
                            model.AddBoolOr([exam_at_day.Not(), exam_at_slot.Not()]).OnlyEnforceIf(exam_at_time.Not())

                            # Now combine with room assignment
                            # If exam assigned to room AND scheduled at this time:
                            assigned_and_scheduled = model.NewBoolVar(f'{exam}_in_{room}_at_{d}_{s}')
                            model.AddBoolAnd([exam_room[(exam, room)], exam_at_time]).OnlyEnforceIf(assigned_and_scheduled)
                            model.AddBoolOr([exam_room[(exam, room)].Not(), exam_at_time.Not()]).OnlyEnforceIf(assigned_and_scheduled.Not())

                            exams_in_room_time.append(assigned_and_scheduled)

                    # Add AtMostOne constraint: only one exam can be assigned to this room at this time
                    model.AddAtMostOne(exams_in_room_time)



        # Ensure computer exams are assigned to a computer room
        for exam in exams: #1 Loop exams
            if exam_types[exam] == "PC": #Check exam
                for room in rooms:
                    uses = rooms[room][0]  

                    if "Computer" not in uses:#3 Check if room can be used as a computer room
                        #4 add constraint
                        model.Add(exam_room[(exam, room)] == 0)


        room_surplus = [] #1 Initialize list of surplus
        for exam in exams:#2 Loop through exams    
            #3 Add constraint the each exam has more than 1 room
            model.Add(sum(exam_room[(exam, room)] for room in rooms) >= 1)
            
            #4Create integer for amount of rooms
            rooms_len = model.NewIntVar(0, 9, f'rooms for {exam}')
            
            model.Add(rooms_len == sum(exam_room[(exam, room)]for room in rooms))


            #5 Create penalty variable
            rooms_penalty = model.NewIntVar(0, 15, f'{exam}_room_surplus_penalty')

            #6 Create Boolean conditions
            is_room_length_greater_6 = model.NewBoolVar(f'{exam}_has_six_or_more_rooms')
            is_room_length_5 = model.NewBoolVar(f'{exam}_has_five_rooms')
            is_room_length_4 = model.NewBoolVar(f'{exam}_has_four_rooms')
            is_room_length_3 = model.NewBoolVar(f'{exam}_has_three_rooms')
            
                #7 Set the true condition
            model.Add(rooms_len >= 6).OnlyEnforceIf(is_room_length_greater_6)
            model.Add(rooms_len <= 5).OnlyEnforceIf(is_room_length_greater_6.Not())
                
            model.Add(rooms_len == 5).OnlyEnforceIf(is_room_length_5)
            model.Add(rooms_len != 5).OnlyEnforceIf(is_room_length_5.Not())

            model.Add(rooms_len == 4).OnlyEnforceIf(is_room_length_4)
            model.Add(rooms_len != 4).OnlyEnforceIf(is_room_length_4.Not())
                
            model.Add(rooms_len == 3).OnlyEnforceIf(is_room_length_3)
            model.Add(rooms_len != 3).OnlyEnforceIf(is_room_length_3.Not())



            #8 Assign penalty values based gap
            model.Add(rooms_penalty == 15).OnlyEnforceIf(is_room_length_greater_6)
            model.Add(rooms_penalty == 9).OnlyEnforceIf(is_room_length_5)
            model.Add(rooms_penalty == 6).OnlyEnforceIf(is_room_length_4)
            model.Add(rooms_penalty == 4).OnlyEnforceIf(is_room_length_3)


            #9 no penalty if less than 3 rooms 
            model.Add(rooms_penalty == 0).OnlyEnforceIf(
                        is_room_length_3.Not(), is_room_length_4.Not(), is_room_length_5.Not(), is_room_length_greater_6.Not(),
                    )
                                #10 Add penalty to total penalties
            room_surplus.append(rooms_penalty)


        #Penalise using pc rooms for non pc exams

        non_pc_exam_penalty = []

        #1 Find computer rooms
        computer_rooms = [room for room in rooms if "Computer" in rooms[room][0]]

        #2 Loop through exams
        for exam in exams:
            #3 if not a PC exams
            if exam_types[exam] != "PC":
                    #4 Check each computer room
                for room in computer_rooms:
                    #5 Create a Boolean 
                    penalty_var = model.NewBoolVar(f"non_pc_exam_in_pc_room_{exam}_{room}")
                    
                    #6 Assign penalty 
                    model.Add(exam_room[(exam, room)] == 1).OnlyEnforceIf(penalty_var)
                    model.Add(exam_room[(exam, room)] != 1).OnlyEnforceIf(penalty_var.Not())

                    #7 Add penality
                    non_pc_exam_penalty.append(5 * penalty_var)




        # Solve the model to minimize the penalties and create a timetable dictionary

        # In[111]:


        model.Minimize(sum( spread_penalties +   extra_time_25_penalties + room_surplus+soft_day_penalties + non_pc_exam_penalty))
        sorted_solutions = []

        class ExamScheduleCollector(cp_model.CpSolverSolutionCallback):
            def __init__(self, exam_day, exam_slot, exam_room, exams, rooms, leader_courses, days, slots, max_solutions=10):
                cp_model.CpSolverSolutionCallback.__init__(self)
                self.exam_day = exam_day
                self.exam_slot = exam_slot
                self.exam_room = exam_room
                self.exams = exams
                self.rooms = rooms
                self.leader_courses = leader_courses
                self.days = days
                self.slots = slots
                self.spread_penalties = spread_penalties or []
                self.soft_day_penalties = soft_day_penalties or []
                self.room_surplus = room_surplus or []
                self.solutions = []
                self.max_solutions = max_solutions

            def on_solution_callback(self):
                schedule = {}
                for exam in self.exams:
                    d = self.Value(self.exam_day[exam])
                    s = self.Value(self.exam_slot[exam])
                    assigned_rooms = [room for room in self.rooms if self.Value(self.exam_room[(exam, room)]) == 1]
                    try:
                        leader = [name for name, exams in self.leader_courses.items() if exam in exams][0]
                    except IndexError:
                        leader = "unknown"
                    schedule[exam] = (d, s, assigned_rooms)

                # Calculate total penalty
                for penalty in [self.spread_penalties, self.soft_day_penalties, self.room_surplus]:
                    for V in penalty:
                        total=self.Value(V)
                    print(f'penalty: {total}')
                total_penalty = sum(self.Value(v) for v in self.spread_penalties + self.soft_day_penalties + self.room_surplus )

                # Store as (schedule, penalty)
                self.solutions.append((schedule, total_penalty))

                if len(self.solutions) >= self.max_solutions:
                    self.StopSearch()

        # Usage:

        solver = cp_model.CpSolver()
        solver.parameters.enumerate_all_solutions = True

        collector = ExamScheduleCollector(
            exam_day, exam_slot, exam_room,
            exams, rooms, leader_courses, days, slots,
            max_solutions=5
        )

        status = solver.Solve(model, collector)

        if status == cp_model.INFEASIBLE:
            print('Infeasible model.')
        elif len(collector.solutions) == 0:
            print('No solutions found.')
            print(status)
        else:
            n_sols = len(collector.solutions)
            print(f"Found {n_sols} solutions.")

            # Sort by penalty (ascending)
            sorted_solutions = sorted(collector.solutions, key=lambda tup: tup[1])  # tup = (schedule, penalty)
            for solution in sorted_solutions:
                schedule, penalty = solution
                print(f"Penalty: {penalty}")



        # This commented out code is if you only want to produce one timetable rather than the numerous ones currenty produced

        # In[112]:


        # model.Minimize(sum(spread_penalties + soft_penalties+unuseds))


        # # Solve the model
        # solver = cp_model.CpSolver()
        # status = solver.Solve(model)

        # if status == cp_model.FEASIBLE or status == cp_model.OPTIMAL:
        #     print("Exam Schedule:")
        #     exams_timetabled = {}
        #     for exam in exams:
        #         print(exam)
        #         d = solver.Value(exam_day[exam])
        #         s = solver.Value(exam_slot[exam])
        #         assigned_rooms = [room for room in rooms if solver.Value(exam_room[(exam, room)]) == 1]
        #         try:
        #             leader = [name for name, exams in leader_courses.items() if exam in exams][0]
        #         except IndexError:
        #             leader = "unknown"
        #         print(f" - {exam} (Leader: {leader}): {days[d]} {slots[s]}")
        #         print(d)
        #         print(s)
        #         exams_timetabled[exam] = (d, s, assigned_rooms)
        # elif status == cp_model.INFEASIBLE:
        #     # print infeasible boolean variables index
        #     print('Infeasible model. Exam schedule could not be created.')
        # else:
        #     print("No solution found.")


        # Presenting the data in an excel spreadsheet

        # In[114]:

        self.select_output_folder()
        for i, (exams_timetabled,Penalties),  in enumerate(sorted_solutions):
        # ------------ BUILD data dictionary ------------
        # data[day][slot] = list of (exam_name, rooms)
            data = {}
            for exam, (d, s, room) in exams_timetabled.items():
                day = days[d]
                slot = s  # keep slot as int 0 or 1
                data.setdefault(day, {}).setdefault(slot, []).append((exam, room))

            # ------------ BUILD rows and row_meta ------------
            rows = []
            row_meta = []  # will store tuples (day_idx, slot_idx) to track rows for merges and coloring

            for d_idx, day_name in enumerate(days):
                for s_idx, slot_name in enumerate(['Morning', 'Afternoon']):
                    exams_list = data.get(day_name, {}).get(s_idx, [])
                    if not exams_list:
                                # No exams this slot — add empty row with '' students and empty room
                        rows.append([day_name, slot_name, '', '', ''])
                        row_meta.append((d_idx, s_idx))
                    else:
                        for exam_name, room in exams_list:
                            room_str = ', '.join(room)
                                    #   total_students = sum(exam_counts.get(exam_name, [0, 0]))  
                            total_students = f'AEA {exam_counts[exam_name][0]}, Non-AEA {exam_counts[exam_name][1]}'
                            if exam_types[exam_name] == "PC":
                                type_str = " (Computer)"
                            elif exam_types[exam_name] == "Standard":
                                type_str = " (Standard)"
                            rows.append([day_name, slot_name, exam_name, total_students, room_str,type_str])
                            row_meta.append((d_idx, s_idx))
                        rows.append([day_name, slot_name, '', '', ''])
                        row_meta.append((d_idx, s_idx))
                    

            # ------------ SAVE to Excel ------------
            df = pd.DataFrame(rows, columns=['Date', 'Time', 'Exam', 'Total No of Students', 'Room', 'Type'])
            filename = f'exam_schedule_merged{i}.xlsx'
            df.to_excel(filename, index=False)

                # ------------ LOAD workbook and worksheet ------------
            wb = load_workbook(filename)
            ws = wb.active

            # ------------ FUNCTION to merge vertical cells ------------
            def merge_vertical(col, key_fn):
                start = 2
                last_key = key_fn(start)
                for r in range(3, ws.max_row + 2):
                    key = key_fn(r) if r <= ws.max_row else None
                    if key != last_key:
                        if r - start > 1:
                            ws.merge_cells(start_row=start, start_column=col,
                                        end_row=r-1, end_column=col)
                        start = r
                        last_key = key

            # Merge Time cells: merge vertically for consecutive identical (Date, Time) pairs
            merge_vertical(2, lambda r: (ws.cell(r,1).value, ws.cell(r,2).value))
                # Merge Date cells: merge vertically across all rows for that day
            merge_vertical(1, lambda r: ws.cell(r, 1).value)



            # ------------ DEFINE fills ------------
            yellow = PatternFill('solid', fgColor='FFFF54')  # bright yellow for Fixed modules
            red = PatternFill('solid', fgColor='EA3323')     # red-orange for Core modules
            blue = PatternFill('solid', fgColor='E0EAF6')    # light blue for alternating rows
            green = PatternFill('solid', fgColor='CBE9B8')   # light green for alternating rows

            # ------------ APPLY alternating row fills BY DAY ------------
            for excel_row, (d_idx, s_idx) in enumerate(row_meta, start=2):  # Excel rows start at 2 (after header)
                fill = blue if d_idx % 2 == 0 else green
                for col in range(1, 6):  # columns A(1) to E(5)
                    ws.cell(row=excel_row, column=col).fill = fill

            # ------------ APPLY fixed/core exam coloring (overwrites cols 3-5) ------------
            for r in range(2, ws.max_row + 1):
                exam_name = ws.cell(r, 3).value
                fill = None
                if exam_name:
                    # Check fixed modules (yellow)
                    if any(exam_name.startswith(fm) for fm in Fixed_modules):
                        fill = yellow
                    # Check core modules (red) overrides yellow
                    if any(exam_name.startswith(cm) for cm in Core_modules):
                        fill = red
                if fill:
                    for c in (3, 4, 5):  # Exam, Total No Students, Room columns
                        ws.cell(r, c).fill = fill
            #Centre text

            for row in range(2, ws.max_row + 1):
                for col in [1, 2]:
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(vertical='center')
                    
            for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter  # Get the column letter (like 'A')

                    for cell in col:
                        try:
                            # Convert cell value to string and get length
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass

                    # Set the column width (add a little extra for padding)
                    ws.column_dimensions[col_letter].width = max_length + 2
            # ------------ SAVE workbook ------------

            output_path = os.path.join(self.output_folder, filename)
            wb.save(output_path)
            print(f"Excel file '{filename}' created with merged cells, colors, and full schedule.")



        # Extract the information from an excel spreadsheet

        # In[115]:


        solutions = []
        def file_reading(filepath, days, slots):
            df = pd.read_excel(filepath)
            exams_timetabled = {}

            for _, row in df.iterrows():
                exam_name = row['Exam']


                day_name = day_name if pd.isna(row['Date']) else row['Date']
                slot_name = slot_name if pd.isna(row['Time']) else (0 if row['Time'] == "Morning" else 1)
                if pd.isna(exam_name) or exam_name == '':
                    continue  # Skip empty rows
                room = row['Room'].split(', ') if pd.notna(row['Room']) and row['Room'] else []

                try:
                    d = days.index(day_name)
                    s = slots.index(slot_name)
                except ValueError:
                    raise ValueError(f"Unrecognized day or slot in file: {day_name} / {slot_name}")

                exams_timetabled[exam_name] = (d, s, room)

            return exams_timetabled
        self.select_test_files()
        for file in self.test_files:
            solutions.append(file_reading(file,days,slots))
        # Checking this schedule has no errors

        # In[118]:


        def get_full_schedule(exams_timetabled, Fixed_modules):
                full_schedule = Fixed_modules.copy()
                full_schedule.update(exams_timetabled)
                return full_schedule


        def check_exam_constraints(student_exams, exams_timetabled, Fixed_modules, Core_modules, module_leaders, extra_time_students_50, exams):
            violations = []
            schedule = get_full_schedule(exams_timetabled,Fixed_modules)

            for exam in exams:
                if exam not in schedule:
                    violations.append(f"❌ Exam '{exam}' is not scheduled in the timetable.")
            
            # 0. Students can't have two exams at the same time
            for student, exs in student_exams.items():
                for i in range(len(exs)):
                    for j in range(i + 1, len(exs)):
                        exam1 = exs[i]
                        exam2 = exs[j]
                        if exams_timetabled[exam1][0] == exams_timetabled[exam2][0] and exams_timetabled[exam1][1] == exams_timetabled[exam2][1]:
                            violations.append(
                                f"❌ Student {student} has two exams '{exam1}' and '{exam2}' at the same time"
                            )
            

            # 1. Core modules fixed: students cannot have more than one core exam on the same day
            for student, exs in student_exams.items():
                core_mods = [exam for exam in exs if exam in Core_modules]
                other_mods = [exam for exam in exs if exam not in Core_modules]

                for core_exam in core_mods:
                    core_day = exams_timetabled[core_exam][0]  # Assume (day, slot, rooms)

                    for other_exam in other_mods:
                        other_day = exams_timetabled[other_exam][0]

                        if core_day == other_day:
                            violations.append(
                                f"❌ Student {student} has core exam '{core_exam}' and non-core exam '{other_exam}' on the same day ({core_day})"
                                )

            # 2. Other modules fixed in date/time (Fixed_modules) 
            for exam, fixed_slot in Fixed_modules.items():
                scheduled_slot = [exams_timetabled.get(exam)[0] , exams_timetabled.get(exam)[1]]
                if scheduled_slot != fixed_slot:
                    violations.append(f"❌ Fixed module '{exam}' is not at the correct time (expected {fixed_slot}, got {scheduled_slot}).")


        # 3. No more than 3 exams in any 2 consecutive days (per student)
            for student, exs in student_exams.items():
                day_count = defaultdict(int)
                for exam in exs:
                    if exam in schedule:
                        day = schedule[exam][0]
                        day_count[day] += 1

                days = sorted(day_count.keys())
                for day in days:
                    next_day = day + 1
                    if next_day in day_count:
                        total = day_count[day] + day_count[next_day]
                        if total > 3:
                            violations.append(
                                f"❌ Student {student} has more than 3 exams across days {day} and {next_day}"
                            )

            # 4. No more than 4 exams in any 5 consecutive weekdays (Monday to Friday)
            for student, exs in student_exams.items():
                day_count = defaultdict(int)
                for exam in exs:
                    if exam in schedule:
                        day = schedule[exam][0]
                        day_count[day] += 1

            all_days = sorted(day_count.keys())
            if all_days:
                min_day, max_day = all_days[0], all_days[-1]
                # Slide over every possible consecutive 5-day window in the exam period
                for start_day in range(min_day, max_day - 4 + 1):
                    total = sum(day_count.get(day, 0) for day in range(start_day, start_day + 5))
                    if total > 4:
                        violations.append(
                            f"❌ Student {student} has more than 4 exams from day {start_day} to {start_day + 4}"
                        )





            # 5. Module leaders cannot have more than one exam in the third week (days 15 to 20 inclusive)
            week3_days = set(range(15, 21))
            for leader, mods in module_leaders.items():
                exams_in_week3 = [exam for exam in mods if exam in schedule and schedule[exam][0] in week3_days]
                if len(exams_in_week3) > 1:
                    violations.append(f"❌ Module leader {leader} has more than one exam in week 3: {exams_in_week3}")

            # 6. Students with >50% extra time cannot have more than one exam on the same day
            for student in extra_time_students_50:
                if student not in student_exams:
                    continue
                day_count = defaultdict(int)
                for exam in student_exams[student]:
                    if exam in schedule:
                        day = schedule[exam][0]
                        day_count[day] += 1
                for day, count in day_count.items():
                    if count > 1:
                        violations.append(f"❌ Student {student} with >50% extra time has {count} exams on day {day}")

            #7 soft Students with 25% extra time cannot have more than one exam on the same day
            for student in AEA:
                if student not in extra_time_students_50:
                    day_count = defaultdict(int)
                    for exam in student_exams[student]:
                        if exam in schedule:
                            day = schedule[exam][0]
                            day_count[day] += 1
                    for day, count in day_count.items():
                        if count > 1:
                            violations.append(f"⚠️soft warning Student {student} with <=25% extra time has {count} exams on day {day}")
                

            #Soft checking theres not more than two exams in any slot in the first week 
            exam_in_slot = defaultdict(list)

            for exam in exams:
                day, slot,rooms = schedule[exam]

                if day <= 15:  # First two weeks
                    exam_in_slot[(day, slot)].append(exam)

            # Check for violations
            for date_slot, scheduled_exams in exam_in_slot.items():
                if len(scheduled_exams) >= 3:
                    violations.append(
                        f"⚠️ Soft warning: day/slot {date_slot} has {len(scheduled_exams)} exams scheduled: {scheduled_exams}"
                    )
            return violations



        def check_room_constraints(
            exams_timetabled,      # dict: exam -> (day, slot, [assigned_rooms])
            exam_counts,           # dict: exam -> (AEA_students, SEQ_students)
            room_dict              # dict: room_name -> [list of types, capacity]
        ):
            violations = []
            # 1. Check room capacity sufficiency per exam
            for exam, (day, slot, rooms) in exams_timetabled.items():
                if exam not in exam_counts:
                    violations.append(f"⚠️ No student count for exam '{exam}', skipping capacity check")
                    continue

                AEA_students, SEQ_students = exam_counts[exam]
                AEA_capacity = sum(room_dict[r][1] for r in rooms if "AEA" in room_dict[r][0])
                SEQ_capacity = sum(room_dict[r][1] for r in rooms if "SEQ" in room_dict[r][0])
                if AEA_capacity < AEA_students:
                    violations.append(
                        f"❌ Exam '{exam}' has insufficient AEA capacity: needed {AEA_students}, assigned {AEA_capacity}"
                    )
                if SEQ_capacity < SEQ_students:
                    violations.append(
                        f"❌ Exam '{exam}' has insufficient SEQ capacity: needed {SEQ_students}, assigned {SEQ_capacity}"
                    )

            # 2. No room double-booked at same day & slot
            room_schedule = defaultdict(list)  # key=(day, slot, room), value=list of exams

            for exam, (day, slot, rooms_) in exams_timetabled.items():
                for room in rooms_:
                    room_schedule[(day, slot, room)].append(exam)


            for (day, slot, room), exams_in_room in room_schedule.items():
                if room != 'NON ME N/A':  # Skip the N/A room 
                    if len(exams_in_room) > 1:
                        violations.append(
                            f"❌ Room '{room}' double-booked on day {day}, slot {slot} for exams: {exams_in_room}"
                        )
                
            # 3. Check computer-based exams are in computer rooms
            for exam, (day, slot, rooms) in exams_timetabled.items():
                if exam_types[exam] == "PC":  # Only check computer-based exams
                    for room in rooms:
                        if "Computer" not in room_dict[room][0]:
                            violations.append(
                                f"❌ Computer-based exam '{exam}' assigned to non-computer room '{room}'"
                            )
            
            # 4 Check every exam assigned at least one room
            for exam, (day, slot, rooms) in exams_timetabled.items():
                if not rooms:
                    violations.append(f"❌ Exam '{exam}' has no assigned room!")

            # 5 Check non PC exams are not in PC rooms
            for exam, (day, slot, rooms) in exams_timetabled.items():
                if exam_types[exam] != "PC":  # Only check non computer-based exams
                    for room in rooms:
                        if "Computer" in room_dict[room][0]:
                            violations.append(
                                f"⚠️ Soft warning: '{exam}' assigned to computer room '{room}' and is not a computer exam"
                            )

            return violations

        for i, exams_timetabled in enumerate(solutions):
        
            violations = check_exam_constraints(
                student_exams=student_exams,
                exams_timetabled=exams_timetabled,
                Fixed_modules=Fixed_modules,
                Core_modules=Core_modules,
                module_leaders=leader_courses,
                extra_time_students_50=extra_time_students_50,
                exams = exams,
        )

            violations.extend(check_room_constraints(
                exams_timetabled=exams_timetabled,
                exam_counts=exam_counts,
                room_dict=rooms
        ))

            if violations:
                for v in violations:
                    print(i)
                    print(v)
            else:
                print("✅ All constraints satisfied!")



if __name__ == "__main__":
    root = tk.Tk()
    app = ExamTimetableUI(root)
    root.mainloop()