# %% [markdown]
# Going to use pandas to read the excel files

# %%
# %pip install pandas
# %pip install rapidfuzz
# %pip install collections
# %pip install openpyxl

# %%
import pandas as pd
from rapidfuzz import process, fuzz
from collections import defaultdict
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re
from dateutil.parser import parse

# %% [markdown]
# Read the files

# %%
students_df = pd.read_excel('/Users/edwardbrady/Library/CloudStorage/OneDrive-ImperialCollegeLondon/UROP-Exam timetabling/input data/student list DONOT SORT ONLY FILTER.xlsx',header=None)
leaders_df = pd.read_excel('/Users/edwardbrady/Library/CloudStorage/OneDrive-ImperialCollegeLondon/UROP-Exam timetabling/input data/module list 2024-25.xlsx',sheet_name=1,header=1)
# Load workbook and worksheet for useful dates
wb = load_workbook("/Users/edwardbrady/Library/CloudStorage/OneDrive-ImperialCollegeLondon/UROP-Exam timetabling/input data/2025-26 Useful Dates.xlsx")
ws = wb.active

# %% [markdown]
# Creating a list of exam names, from excel can see they're in Row 'A' which is 0 and start from column 'J' which is 9

# %%
exam_names = students_df.iloc[0, 9:].dropna().tolist()
print(exam_names)

# %% [markdown]
# To calculate the number of days for the exam period we can assume it will be 3 weeks, weeks 31-33 and therefore it will be 21 days starting on monday of week 31

# %%
num_days = 21

# %% [markdown]
# To create a dictionary of students and there exams we will flick through each row, which is a studnet, and for every module that has an x , A or B add it to their exams. At the end add PEN as everyone does PEN
# 

# %%
# Get the range of rows containing student data (from row 3 onward)
student_rows = students_df.iloc[2:, :]  # row index 3 and onward

# Initialize the dictionary
student_exams = {}

for _, row in student_rows.iterrows():
    cid = row[0]  # Column A = student CID
    exams_taken = []

    for col_idx, exam_name in enumerate(exam_names, start=9):  # Column J = index 9
        if str(row[col_idx]).strip().lower() == 'x':  # Check for 'x' (case-insensitive)
            exams_taken.append(exam_name)

    #Add in PEN exam
    exams_taken.append('MECH60015 / MECH70030 Professional Engineering Skills (ME3/AME) (ECE exam)')
    student_exams[cid] = exams_taken



print(student_exams)

# %% [markdown]
# To create a list of core modules I will read it off the excel from last year 

# %%
Core_modules = ["MECH70001 Nuclear Thermal Hydraulics","MECH60004/MECH70042 Introduction to Nuclear Energy A/B","MECH70002 Nuclear Reactor Physics","MECH70008 Mechanical Transmissions Technology","MECH70006 Metal Processing Technology","MECH70021Aircraft Engine Technology","MECH70003 Future Clean Transport Technology","MECH60015 / MECH70030 Professional Engineering Skills (ME3/AME) (ECE exam)"]

# %% [markdown]
# Creating a dictionary of fixed modules and their dates, in the format day, slot with day being the day from the first monday (including w/e) of exam season and slot being either morning or afternoon

# %%
Fixed_modules = {"BUSI60039 Business Strategy" :[1,1],"BUSI60046 Project Management":[2,1],"ME-ELEC70098 Optimisation":[3,0],"MECH70001 Nuclear Thermal Hydraulics":[3,0],"BUSI60040/BUSI60043 Corporate Finance Online/Finance & Financial Management":[3,1],"MECH60004/MECH70042 Introduction to Nuclear Energy A/B":[4,0],"ME-ELEC70022 Modelling and Control of Multi-body Mechanical Systems":[4,0],"MATE97022 Nuclear Materials 1":[4,0],"ME-MATE70029 Nuclear Fusion":[9,0],"MECH70002 Nuclear Reactor Physics":[10,0],"ME-ELEC70076 Sustainable Electrical Systems":[10,0],"ME ELEC70066 Applied Advanced Optimisation":[10,0],"MECH70020 Combustion, Safety and Fire Dynamics":[11,0],"BIOE70016 Human Neuromechanical Control and Learning":[11,0],"CENG60013 Nuclear Chemical Engineering":[11,0],"MECH70008 Mechanical Transmissions Technology":[13,1],"MECH70006 Metal Processing Technology":[13,1],"MECH70021Aircraft Engine Technology":[13,1],"MECH70003 Future Clean Transport Technology":[13,1],"MECH60015 / MECH70030 Professional Engineering Skills (ME3/AME) (ECE exam)":[14,1]}

# %% [markdown]
# Creat a dictionary of each module leader and there respective exams, this needs to use the names used in the students spreadsheet and as such we combine course code and name and find the closest match. Some are disregarded as they're not examinable courses.

# %%

# Extract official module names from row 0, columns J onwards (i.e., column 9 onward, 0-indexed)
standardized_names = students_df.iloc[0, 9:].dropna().astype(str).str.strip().tolist()


# Prepare module-leader dictionary
leader_courses = defaultdict(list)

# Loop through rows in the module list
for _, row in leaders_df.iterrows():
    leader = row['Module Leader (lecturer 1)']
    name = row['Module Name']
    code = row['Banner Code (New CR)']   # module leader

    # Skip if any required field is missing
    if pd.isna(code) or pd.isna(name) or pd.isna(leader):
        continue

    if leader == "n/a":
        continue

    # Combine code and name
    combined_name = f"{code} {name}"

    # Fuzzy match to standardized names
    best_match, score, _ = process.extractOne(
        combined_name, standardized_names, scorer=fuzz.token_sort_ratio
    )

    if score >= 70:
        if best_match not in leader_courses[leader]:
            leader_courses[leader].append(best_match)
        else:
            print(f"⚠️ Duplicate match skipped for '{combined_name}': '{best_match}' is already listed for {leader}.")
    else:
        print(f"⚠️ Low confidence match for '{combined_name}' (best: '{best_match}', score: {score}).")


# Convert to normal dict if desired
leader_courses = dict(leader_courses)

print("Module leaders and their courses:")
print(leader_courses)

# %% [markdown]
# Make a list of students with 25% extra time to ensure they dont have more than on exam a day

# %%
extra_time_students_25 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("15min/hour", "25% extra time"))].iloc[:, 0].tolist()

print(extra_time_students_25)

# %%
extra_time_students_50 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("30min/hour", "50% extra time"))].iloc[:, 0].tolist()

print(extra_time_students_50)

# %% [markdown]
# Need to read the bank holidays and forbidden time slots these will have the form [day,slot] the weekends, (days 5, 6 and 12, 13) can be assumed to not have exams and the no exam on morning of the last friday

# %%
no_exam_dates = [[5,0],[5,1],[6,0],[6,1],[12,0],[12,1],[13,0],[13,1],[20,0],]

# %% [markdown]
# To find the bank holidays we will use the useful dates spreadsheet as this has the start of the summer term and also the bank holidays, first we find the date of the start of summer term, this is made more difficult due to the fact that it is written in the form of day, month to day month year and we need day month year of start. Once this is found the first monday is assumed to be the following monday from this. Then the dates of the bank holidays are found and anything within 20 days is found and appended to the no exam dates

# %%


# Initialize
bank_holidays = []


# --- Step 1: Extract bank holidays from col F (names) and G (dates) ---
row = 5
while True:
    name = ws[f"F{row}"].value
    date_cell = ws[f"G{row}"].value
    if name is None or "Term Dates" in str(name):
        break
    if isinstance(date_cell, datetime):
        bank_holidays.append((str(name).strip(), date_cell.date()))
    row += 1

# --- Step 2: Find Summer Term start date from section below ---
summer_start = None
while row < ws.max_row:
    cell_value = ws[f"F{row}"].value
    if cell_value and "Summer Term" in str(cell_value):
        term_range = ws[f"F{row + 1}"].value
        if term_range:
            try:
                # Extract left side of range before "to", drop weekday (e.g., "Fri"), and append year
                start_part = term_range.split("to")[0].strip()
                start_str = re.sub(r"^\w+\s+", "", start_part)  # Removes "Fri", leaves "24 Apr"
                # Try extracting year from second part if present
                year_match = re.search(r"\b\d{4}\b", term_range)
                if year_match:
                    start_str += f" {year_match.group(0)}"
                else:
                    raise ValueError("Year not found in date range.")
                summer_start = parse(start_str, dayfirst=True).date()
            except Exception as e:
                raise ValueError(f"Could not parse Summer Term start: {term_range}") from e
        else:
            raise ValueError("Summer Term range cell is empty.")
        break
    row += 1

if not summer_start:
    raise ValueError("Summer Term start date not found.")

# --- Step 3: Find first Monday on or after summer_start ---
first_monday = summer_start
while first_monday.weekday() != 0:  # 0 = Monday
    first_monday += timedelta(days=1)

# --- Step 4: Find bank holidays within 3 weeks after first Monday ---
for name, bh_date in bank_holidays:
    delta = (bh_date - first_monday).days
    if 0 <= delta <= 20:
        print(f"{name} is {delta} days after first Monday ({bh_date})")
        no_exam_dates.append([delta, 0])
        no_exam_dates.append([delta, 1])

# --- Final Output ---
print("No exam dates:", no_exam_dates)



# %% [markdown]
# # Exam Timetabling with Morning and Afternoon Slots using OR-Tools
# 
# This notebook models exams scheduled over multiple days with morning and afternoon slots.
# 
# Constraints:
# - No student can have two exams in the same slot (same day & same slot)
# - No more than 2 exams per slot (day + morning/afternoon)
# 
# Outputs the exam schedule showing day and slot assignment.

# %%
# Install OR-Tools if you haven't yet
#!pip install ortools

# %%
from ortools.sat.python import cp_model

# %%


# %%
model = cp_model.CpModel()


num_slots = len(slots)

# Variables: exam_day and exam_slot
exam_day = {}
exam_slot = {}
for exam in exams:
    exam_day[exam] = model.NewIntVar(0, num_days - 1, f'{exam}_day')
    exam_slot[exam] = model.NewIntVar(0, num_slots - 1, f'{exam}_slot')

# %%

for student in student_exams:
    for d in range(num_days - 1):  #1 Go through each pair of 2 days
        exams_in_2_days = []

        for exam in student_exams[student]:  # Go through their exams
		        #2 Create the Boolean variables
            is_on_d = model.NewBoolVar(f'{student}_{exam}_on_day_{d}')
            is_on_d1 = model.NewBoolVar(f'{student}_{exam}_on_day_{d+1}')
            is_in_either = model.NewBoolVar(f'{student}_{exam}_on_day_{d}_or_{d+1}')

            #3 These say whether an exam is on day d or d+1
            model.Add(exam_day[exam] == d).OnlyEnforceIf(is_on_d)
            model.Add(exam_day[exam] != d).OnlyEnforceIf(is_on_d.Not())
            model.Add(exam_day[exam] == d+1).OnlyEnforceIf(is_on_d1)
            model.Add(exam_day[exam] != d+1).OnlyEnforceIf(is_on_d1.Not())

            #4 This says: if the exam is on either day, mark it as true
            model.AddBoolOr([is_on_d, is_on_d1]).OnlyEnforceIf(is_in_either)
            model.AddBoolAnd([is_on_d.Not(), is_on_d1.Not()]).OnlyEnforceIf(is_in_either.Not())

            exams_in_2_days.append(is_in_either)

        #5 Final rule: no more than 3 exams in this 2-day window
        model.Add(sum(exams_in_2_days) <= 3)

for student in student_exams:
    for start_day in range(num_days - 4):  #1 sliding window
        exams_in_window = []
        for exam in student_exams[student]: #2 Each exam a student has
		        #3 Add in window variable
            in_window = model.NewBoolVar(f'{student}_{exam}_in_day_{start_day}_to_{start_day+4}')
            #4 Check whether exam in window
            model.AddLinearConstraint(exam_day[exam], start_day, start_day + 4).OnlyEnforceIf(in_window)
            model.AddBoolOr([exam_day[exam] < start_day, exam_day[exam] > start_day + 4]).OnlyEnforceIf(in_window.Not())
            #5 Add exam to list
            exams_in_window.append(in_window)
        #6 Add constraint
        model.Add(sum(exams_in_window) <= 4)

for student, exams in student_exams.items():#1 Loop students
		#2 Create a list of their core exams
    core_mods = [exam for exam in exams if exam in Core_modules]
    other_mods = [exam for exam in exams if exam not in Core_modules]
		#3 Loop through the core exams
    for exam in core_mods:
        for other in other_mods:
		        #4 Add constraint
            model.Add(exam_day[exam] != exam_day[other])

for exam, fixed_day_slot in Fixed_modules.items():
	model.Add(exam_day[exam] == fixed_day_slot[0])
	model.Add(exam_slot[exam] == fixed_day_slot[1])

for leader in leader_courses: #1 Loop through module leaders
    week_3_exams = []
    for exam in leader_courses[leader]: #2 Loop through each module leaders exams
		    #3 Initialize variable
        is_in_week3 = model.NewBoolVar(f'{exam}_in_week3')
        #4 Check to see if exam in week 3
        model.AddLinearConstraint(exam_day[exam], 13, 20).OnlyEnforceIf(is_in_week3)
        model.AddBoolOr([exam_day[exam] < 13, exam_day[exam] > 20]).OnlyEnforceIf(is_in_week3.Not())
        #5 Add exam to list
        week_3_exams.append(is_in_week3)
    #6 Add constraint
    model.Add(sum(week_3_exams) <= 1)
	
for student in extra_time_students_50:#1 Loop through extra time students
    for day in num_days:#2 Loop through days
        exams_on_day = []

        for exam in student_exams[student]:
		        #3 Initialize variable
            is_on_day = model.NewBoolVar(f'{student}_{exam}_on_day_{day}')
            #4 Check if exam is on day
            model.Add(exam_day[exam] == day).OnlyEnforceIf(is_on_day)
            model.Add(exam_day[exam] != day).OnlyEnforceIf(is_on_day.Not())
						#5 Add exam to list
            exams_on_day.append(is_on_day)

        #6 Add constraint 
        model.Add(sum(exams_on_day) <= 1)
for exam in exams:#1 Loop through exams
    for date, slot in no_exam_dates: #2 Loop through forbidden days and times
        #3 Init boolean variable 
        is_forbidden = model.NewBoolVar(f'{exam}_forbidden_{date}_{slot}')

        #4 Link the boolean to the conditions
        model.Add(exam_day[exam] == date).OnlyEnforceIf(is_forbidden)
        model.Add(exam_day[exam] != date).OnlyEnforceIf(is_forbidden.Not())

        model.Add(exam_slot[exam] == slot).OnlyEnforceIf(is_forbidden)
        model.Add(exam_slot[exam] != slot).OnlyEnforceIf(is_forbidden.Not())

        #5 Add constraint
        model.Add(is_forbidden == 0)
    
soft_penalties = []

for student in extra_time_students_25:#1 Loop through extra time students
    for day in num_days:#2 Loop through days
        exams_on_day = []

        for exam in student_exams[student]:
		        #3 Initialize variable
            is_on_day = model.NewBoolVar(f'{student}_{exam}_on_day_{day}')
            #4 Check if exam is on day
            model.Add(exam_day[exam] == day).OnlyEnforceIf(is_on_day)
            model.Add(exam_day[exam] != day).OnlyEnforceIf(is_on_day.Not())
						#5 Add exam to list
            exams_on_day.append(is_on_day)

        #6 Add constraint 
        penalty = model.NewIntVar(0, 1, f'{student}_penalty_day_{day}')
        model.Add(penalty == 1).OnlyEnforceIf(exams_on_day > 1)
        model.Add(penalty == 0).OnlyEnforceIf(exams_on_day <= 1)

        soft_penalties.append(penalty)

spread_penalties =[]

for leader in leader_courses:#1 Loop through module leaders
    mods = leader_courses[leader]
    
    for i in range(len(mods)):#2 loop through modules
        for j in range(i+1, len(mods)):
            m1 = mods[i]
            m2 = mods[j]

            #3 Calculate absolute day difference
            diff = model.NewIntVar(-21, 21, f'{m1}_{m2}_diff')
            abs_diff = model.NewIntVar(0, 21, f'{m1}_{m2}_abs_diff')
						#4 Add difference to model
            model.Add(diff == exam_day[m1] - exam_day[m2])
            model.AddAbsEquality(abs_diff, diff)

            #5 Create penalty variable
            close_penalty = model.NewIntVar(0, 5, f'{m1}_{m2}_penalty')

            #6 Create Boolean conditions
            is_gap_3 = model.NewBoolVar(f'{m1}_{m2}_gap3')
            is_gap_2 = model.NewBoolVar(f'{m1}_{m2}_gap2')
            is_gap_1 = model.NewBoolVar(f'{m1}_{m2}_gap1')
            is_gap_0 = model.NewBoolVar(f'{m1}_{m2}_gap0')
						#7 Set the true condition
            model.Add(abs_diff == 3).OnlyEnforceIf(is_gap_3)
            model.Add(abs_diff != 3).OnlyEnforceIf(is_gap_3.Not())

            model.Add(abs_diff == 2).OnlyEnforceIf(is_gap_2)
            model.Add(abs_diff != 2).OnlyEnforceIf(is_gap_2.Not())

            model.Add(abs_diff == 1).OnlyEnforceIf(is_gap_1)
            model.Add(abs_diff != 1).OnlyEnforceIf(is_gap_1.Not())

            model.Add(abs_diff == 0).OnlyEnforceIf(is_gap_0)
            model.Add(abs_diff != 0).OnlyEnforceIf(is_gap_0.Not())

            #8 Assign penalty values based gap
            model.Add(close_penalty == 1).OnlyEnforceIf(is_gap_3)
            model.Add(close_penalty == 3).OnlyEnforceIf(is_gap_2)
            model.Add(close_penalty == 4).OnlyEnforceIf(is_gap_1)
            model.Add(close_penalty == 5).OnlyEnforceIf(is_gap_0)

            #9 no penalty if gap ≥ 4 and not equal to 0–3
            model.Add(close_penalty == 0).OnlyEnforceIf(
                is_gap_3.Not(), is_gap_2.Not(), is_gap_1.Not(), is_gap_0.Not()
            )
						#10 Add penalty to total penalties
            spread_penalties.append(close_penalty)

model.Minimize(sum(spread_penalties + soft_penalties))

# %%
# Solve the model
solver = cp_model.CpSolver()
status = solver.Solve(model)

if status == cp_model.FEASIBLE or status == cp_model.OPTIMAL:
    print("Exam Schedule:")
    for exam in exams:
        d = solver.Value(exam_day[exam])
        s = solver.Value(exam_slot[exam])
        leader = leader_courses.get(exam, 'Unknown')
        print(f" - {exam} (Leader: {leader}): {days[d]} {slots[s]}")
else:
    print("No solution found.")


# %%
# %pip install pandas openpyxl

# %%
import pandas as pd
from openpyxl import load_workbook

# %% [markdown]
# Some example data of what an output dictionary could look like with it being exam name and then day then slot (morning/afternoon), this only has 4 days and a few exams


# Group exams by day and slot
data = {}
for exam, (d, s) in exams_timetabled.items():
    day = days[d]
    slot = slots[s]
    data.setdefault(day, {}).setdefault(slot, []).append(exam)

rows = []
for day in days:
    for slot in slots:
        exams_list = data.get(day, {}).get(slot, [])
        exams_str = ', '.join(exams_list) if exams_list else ''
        rows.append([day, slot, exams_str])


df = pd.DataFrame(rows, columns=['Day', 'Time', 'Exams'])

# Save DataFrame to Excel
filename = 'exam_schedule_merged.xlsx'
df.to_excel(filename, index=False)

# Now open with openpyxl to merge cells
wb = load_workbook(filename)
ws = wb.active

def merge_cells_for_col(col_idx):
    current_val = None
    start_row = 2  # skip header
    max_row = ws.max_row + 1
    for row in range(2, max_row):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value != current_val:
            # Merge the previous block if >1 row
            if current_val is not None and row - start_row > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx,
                               end_row=row - 1, end_column=col_idx)
            current_val = cell.value
            start_row = row
    # Merge last block
    if current_val is not None and max_row - start_row > 0:
        ws.merge_cells(start_row=start_row, start_column=col_idx,
                       end_row=max_row - 1, end_column=col_idx)

# Merge Day column (A=1)
merge_cells_for_col(1)
# Merge Time column (B=2)
merge_cells_for_col(2)

wb.save(filename)
print(f"Excel file '{filename}' created with merged cells.")



